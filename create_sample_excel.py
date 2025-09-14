#!/usr/bin/env python3
"""
Create Sample Excel File for Barcode Import Testing
"""
from openpyxl import Workbook

def create_sample_barcode_excel():
    """Create a sample Excel file with barcodes for testing"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Barcodes"
    
    # Add headers
    headers = ['barcode', 'notes']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add sample barcodes (these are real barcodes that should work)
    sample_barcodes = [
        ['015000047757', 'Yogurt blends snack'],
        ['3017620425035', 'Nutella'],
        ['3175680011480', 'Sésame'],
        ['1234567890123', 'Test barcode - should fail'],
        ['8901030865000', 'Parle-G Biscuits'],
        ['8901030865001', 'Another test barcode']
    ]
    
    # Add sample data
    for row, (barcode, notes) in enumerate(sample_barcodes, 2):
        ws.cell(row=row, column=1, value=barcode)
        ws.cell(row=row, column=2, value=notes)
    
    # Save the file
    filename = 'sample_barcodes.xlsx'
    wb.save(filename)
    print(f"✅ Created sample Excel file: {filename}")
    print(f"📊 Contains {len(sample_barcodes)} sample barcodes")
    print(f"📋 Headers: {headers}")
    
    return filename

if __name__ == "__main__":
    create_sample_barcode_excel()
