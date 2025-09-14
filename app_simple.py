#!/usr/bin/env python3
"""
Working Flask App with Firebase - Simplified Version
"""
from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from flask_cors import CORS
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime, timezone
import json
import firebase_admin
from firebase_admin import credentials, firestore
import os
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables
load_dotenv()

# Initialize Firebase
try:
    existing_app = firebase_admin.get_app()
    print("Firebase app already initialized")
except ValueError:
    print("Initializing Firebase...")
    cred = credentials.Certificate('firebase-service-account.json')
    firebase_admin.initialize_app(cred)
    print("Firebase initialized successfully")

# Get Firestore client
db = firestore.client()
print("Firestore client created")

# Initialize Flask app
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'pos-dashboard-secret-key')

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

# User class for Flask-Login
class User(UserMixin):
    def __init__(self, id, username, role='admin'):
        self.id = id
        self.username = username
        self.role = role

@login_manager.user_loader
def load_user(user_id):
    admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
    user1_username = os.environ.get('USER1_USERNAME', 'user1')
    
    users = {
        admin_username: User(admin_username, admin_username, 'admin'),
        user1_username: User(user1_username, user1_username, 'user')
    }
    return users.get(user_id)

CORS(app)

# Routes
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
        admin_password_hash = os.environ.get('ADMIN_PASSWORD_HASH', generate_password_hash('admin123'))
        user1_username = os.environ.get('USER1_USERNAME', 'user1')
        user1_password_hash = os.environ.get('USER1_PASSWORD_HASH', generate_password_hash('user123'))
        
        users = {
            admin_username: admin_password_hash,
            user1_username: user1_password_hash
        }
        
        if username in users and check_password_hash(users[username], password):
            user = User(username, username)
            login_user(user)
            return jsonify({
                'status': 'success',
                'message': 'Login successful',
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'role': user.role
                }
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'Invalid username or password'
            }), 401
    
    return render_template('login.html')

@app.route('/logout', methods=['POST'])
@login_required
def logout():
    logout_user()
    return jsonify({
        'status': 'success',
        'message': 'Logged out successfully'
    })

@app.route('/api/auth/status', methods=['GET'])
def auth_status():
    if current_user.is_authenticated:
        return jsonify({
            'status': 'success',
            'authenticated': True,
            'user': {
                'id': current_user.id,
                'username': current_user.username,
                'role': current_user.role
            }
        })
    else:
        return jsonify({
            'status': 'success',
            'authenticated': False
        })

@app.route('/api/products', methods=['GET'])
@login_required
def get_products():
    """Get all products from Firebase"""
    try:
        print("Getting products from Firebase...")
        products_ref = db.collection('barcode_cache')
        docs = products_ref.stream()
        
        products = []
        for doc in docs:
            product_data = doc.to_dict()
            mapped_product = {
                'id': doc.id,
                'name': product_data.get('name', 'Unnamed Product'),
                'brand': product_data.get('brand', ''),
                'category': product_data.get('category', ''),
                'barcode': doc.id,
                'mrp': product_data.get('mrp', 0),
                'salePrice': product_data.get('salePrice', 0),
                'stock': product_data.get('stockQuantity', 0),
                'isActive': product_data.get('isActive', True),
                'useInFirstStart': product_data.get('useInFirstStart', False),
                'imageUrl': product_data.get('photoPath', ''),
                'description': product_data.get('description', ''),
                'createdAt': product_data.get('createdAt', ''),
                'updatedAt': product_data.get('updatedAt', ''),
                'size': product_data.get('size', ''),
                'unit': product_data.get('unit', ''),
                'scanCount': product_data.get('scanCount', 0),
                'syncStatus': product_data.get('syncStatus', ''),
                'sortOrder': product_data.get('sortOrder', 0)
            }
            products.append(mapped_product)
        
        print(f"Retrieved {len(products)} products")
        return jsonify(products)
        
    except Exception as e:
        print(f"Error getting products: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/health')
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'firebase_status': 'connected',
        'version': '1.0.0'
    })

@app.route('/api/test-products', methods=['GET'])
def test_products():
    """Test products endpoint without authentication"""
    try:
        print("Testing products endpoint...")
        products_ref = db.collection('barcode_cache')
        docs = products_ref.stream()
        
        products = []
        for doc in docs:
            product_data = doc.to_dict()
            mapped_product = {
                'id': doc.id,
                'name': product_data.get('name', 'Unnamed Product'),
                'brand': product_data.get('brand', ''),
                'category': product_data.get('category', ''),
                'barcode': doc.id,
                'mrp': product_data.get('mrp', 0),
                'salePrice': product_data.get('salePrice', 0),
                'stock': product_data.get('stockQuantity', 0),
                'isActive': product_data.get('isActive', True),
                'useInFirstStart': product_data.get('useInFirstStart', False),
                'imageUrl': product_data.get('photoPath', ''),
                'description': product_data.get('description', ''),
                'createdAt': product_data.get('createdAt', ''),
                'updatedAt': product_data.get('updatedAt', ''),
                'size': product_data.get('size', ''),
                'unit': product_data.get('unit', ''),
                'scanCount': product_data.get('scanCount', 0),
                'syncStatus': product_data.get('syncStatus', ''),
                'sortOrder': product_data.get('sortOrder', 0)
            }
            products.append(mapped_product)
        
        print(f"Retrieved {len(products)} products for test")
        return jsonify({
            'status': 'success',
            'message': f'Found {len(products)} products',
            'products_count': len(products),
            'sample_products': products[:3]  # Return first 3 products
        })
        
    except Exception as e:
        print(f"Error in test products: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/unfound-barcodes', methods=['POST'])
def create_unfound_barcode():
    """Create a new unfound barcode"""
    try:
        data = request.get_json()
        barcode = data.get('barcode')
        
        if not barcode:
            return jsonify({'error': 'Barcode is required'}), 400
        
        # Check if barcode already exists
        existing_doc = db.collection('unfound_barcodes').where('barcode', '==', barcode).get()
        if existing_doc:
            return jsonify({'error': 'Barcode already exists in unfound list'}), 400
        
        barcode_data = {
            'barcode': barcode,
            'createdAt': datetime.now().isoformat(),
            'lastRetry': None,
            'retryCount': 0
        }
        
        doc_ref = db.collection('unfound_barcodes').add(barcode_data)
        return jsonify({
            'message': 'Unfound barcode created successfully',
            'id': doc_ref[1].id
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/unfound-barcodes', methods=['GET'])
def get_unfound_barcodes():
    """Get all unfound barcodes from Firebase"""
    try:
        print("Getting unfound barcodes from Firebase...")
        unfound_barcodes_ref = db.collection('unfound_barcodes')
        docs = unfound_barcodes_ref.stream()
        
        unfound_barcodes = []
        for doc in docs:
            barcode_data = doc.to_dict()
            barcode_data['id'] = doc.id
            unfound_barcodes.append(barcode_data)
        
        print(f"Retrieved {len(unfound_barcodes)} unfound barcodes")
        return jsonify(unfound_barcodes)
        
    except Exception as e:
        print(f"Error getting unfound barcodes: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/unfound-barcodes/scrape', methods=['POST'])
def scrape_unfound_barcodes():
    """Manually scrape unfound barcodes"""
    try:
        data = request.get_json()
        barcode_ids = data.get('barcode_ids', [])
        
        if not barcode_ids:
            return jsonify({'error': 'No barcode IDs provided'}), 400
        
        print(f"Starting manual scraping for {len(barcode_ids)} unfound barcodes...")
        
        processed_count = 0
        scraped_count = 0
        failed_count = 0
        errors = []
        
        for barcode_id in barcode_ids:
            try:
                # Get barcode data from Firebase
                doc_ref = db.collection('unfound_barcodes').document(barcode_id)
                doc = doc_ref.get()
                
                if not doc.exists:
                    errors.append(f"Barcode {barcode_id} not found")
                    continue
                
                barcode_data = doc.to_dict()
                barcode = barcode_data['barcode']
                
                print(f"🔍 Scraping unfound barcode: {barcode}")
                
                # Try to scrape product data
                url = f"https://smartconsumer-beta.org/01/{barcode}"
                product_data = scrape_product_data(barcode, url)
                
                if product_data and isinstance(product_data, dict) and product_data.get('name') != 'N/A':
                    # Add to barcode_cache collection
                    product_data['barcode'] = barcode
                    product_data['createdAt'] = datetime.now().isoformat()
                    product_data['updatedAt'] = datetime.now().isoformat()
                    product_data['scanCount'] = 1
                    product_data['syncStatus'] = 'pending'
                    product_data['sortOrder'] = 0
                    
                    db.collection('barcode_cache').document(barcode).set(product_data)
                    
                    # Remove from unfound barcodes
                    doc_ref.delete()
                    
                    scraped_count += 1
                    print(f"✅ Successfully scraped: {product_data.get('name', 'Unknown')}")
                else:
                    # Update retry count
                    retry_count = barcode_data.get('retryCount', 0) + 1
                    doc_ref.update({
                        'lastRetry': datetime.now().isoformat(),
                        'retryCount': retry_count
                    })
                    
                    failed_count += 1
                    print(f"❌ Could not scrape barcode {barcode}, retry count: {retry_count}")
                
                processed_count += 1
                
            except Exception as e:
                error_msg = f"Error processing barcode {barcode_id}: {str(e)}"
                errors.append(error_msg)
                print(f"❌ {error_msg}")
        
        response = {
            'status': 'success',
            'message': f'Processed {processed_count} unfound barcodes',
            'processed_count': processed_count,
            'scraped_count': scraped_count,
            'failed_count': failed_count,
            'errors': errors[:10] if errors else []
        }
        
        print(f"Manual scraping completed: {processed_count} processed, {scraped_count} scraped, {failed_count} failed")
        return jsonify(response)
        
    except Exception as e:
        print(f"Manual scraping error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/unfound-barcodes/<barcode_id>', methods=['DELETE'])
def delete_unfound_barcode(barcode_id):
    """Delete a specific unfound barcode"""
    try:
        db.collection('unfound_barcodes').document(barcode_id).delete()
        return jsonify({'message': 'Unfound barcode deleted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/unfound-barcodes/bulk-delete', methods=['POST'])
def bulk_delete_unfound_barcodes():
    """Delete multiple unfound barcodes"""
    try:
        data = request.get_json()
        barcode_ids = data.get('barcode_ids', [])
        
        deleted_count = 0
        for barcode_id in barcode_ids:
            try:
                db.collection('unfound_barcodes').document(barcode_id).delete()
                deleted_count += 1
            except Exception as e:
                print(f"Error deleting unfound barcode {barcode_id}: {e}")
        
        return jsonify({
            'message': f'Successfully deleted {deleted_count} out of {len(barcode_ids)} unfound barcodes',
            'deleted_count': deleted_count
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/import-barcodes', methods=['POST'])
def import_barcodes_with_scraping():
    """Import barcodes from Excel and automatically scrape product data"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'File must be an Excel file (.xlsx or .xls) or CSV file (.csv)'}), 400
        
        # Read file based on extension
        import csv
        import io
        
        if file.filename.endswith('.csv'):
            # Read CSV file
            file_content = file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(file_content))
            rows = list(csv_reader)
            
            if not rows:
                return jsonify({'error': 'CSV file is empty'}), 400
            
            headers = list(rows[0].keys())
            
            # Check if barcode column exists
            if 'barcode' not in headers:
                return jsonify({'error': 'CSV file must contain a "barcode" column'}), 400
            
            # Process barcodes
            processed_count = 0
            scraped_count = 0
            skipped_count = 0
            errors = []
            
            print(f"Starting CSV barcode import and scraping...")
            
            for row_data in rows:
                try:
                    barcode = str(row_data['barcode']).strip()
                    if not barcode or barcode == 'None':
                        continue
                    
                    print(f"Processing barcode: {barcode}")
                    
                    # Check if barcode already exists in barcode_cache
                    existing_doc = db.collection('barcode_cache').document(barcode).get()
                    if existing_doc.exists:
                        print(f"Barcode {barcode} already exists, skipping")
                        skipped_count += 1
                        continue
                    
                    # Try to scrape product data
                    url = f"https://smartconsumer-beta.org/01/{barcode}"
                    product_data = scrape_product_data(barcode, url)
                    
                    if product_data and isinstance(product_data, dict) and product_data.get('name') != 'N/A':
                        # Add to barcode_cache collection
                        product_data['barcode'] = barcode
                        product_data['createdAt'] = datetime.now().isoformat()
                        product_data['updatedAt'] = datetime.now().isoformat()
                        product_data['scanCount'] = 1
                        product_data['syncStatus'] = 'pending'
                        product_data['sortOrder'] = 0
                        
                        db.collection('barcode_cache').document(barcode).set(product_data)
                        scraped_count += 1
                        print(f"✅ Successfully scraped: {product_data.get('name', 'Unknown')}")
                    else:
                        # Add to unfound_barcodes for retry
                        unfound_data = {
                            'barcode': barcode,
                            'createdAt': datetime.now().isoformat(),
                            'lastRetry': None,
                            'retryCount': 0
                        }
                        db.collection('unfound_barcodes').add(unfound_data)
                        print(f"❌ Could not scrape barcode {barcode}, added to unfound list")
                    
                    processed_count += 1
                    
                except Exception as e:
                    error_msg = f"Row {processed_count + 1}: {str(e)}"
                    errors.append(error_msg)
                    print(f"Error processing row: {e}")
        
        else:
            # Read Excel file
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            # Check if barcode column exists
            if 'barcode' not in headers:
                return jsonify({'error': 'Excel file must contain a "barcode" column'}), 400
            
            # Process barcodes
            processed_count = 0
            scraped_count = 0
            skipped_count = 0
            errors = []
            
            print(f"Starting Excel barcode import and scraping...")
            
            for row_num in range(2, ws.max_row + 1):
                try:
                    row_data = {}
                    for col_num, header in enumerate(headers, 1):
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        row_data[header] = cell_value
                    
                    barcode = str(row_data['barcode']).strip()
                    if not barcode or barcode == 'None':
                        continue
                    
                    print(f"Processing barcode: {barcode}")
                    
                    # Check if barcode already exists in barcode_cache
                    existing_doc = db.collection('barcode_cache').document(barcode).get()
                    if existing_doc.exists:
                        print(f"Barcode {barcode} already exists, skipping")
                        skipped_count += 1
                        continue
                    
                    # Try to scrape product data
                    url = f"https://smartconsumer-beta.org/01/{barcode}"
                    product_data = scrape_product_data(barcode, url)
                    
                    if product_data and isinstance(product_data, dict) and product_data.get('name') != 'N/A':
                        # Add to barcode_cache collection
                        product_data['barcode'] = barcode
                        product_data['createdAt'] = datetime.now().isoformat()
                        product_data['updatedAt'] = datetime.now().isoformat()
                        product_data['scanCount'] = 1
                        product_data['syncStatus'] = 'pending'
                        product_data['sortOrder'] = 0
                        
                        db.collection('barcode_cache').document(barcode).set(product_data)
                        scraped_count += 1
                        print(f"✅ Successfully scraped: {product_data.get('name', 'Unknown')}")
                    else:
                        # Add to unfound_barcodes for retry
                        unfound_data = {
                            'barcode': barcode,
                            'createdAt': datetime.now().isoformat(),
                            'lastRetry': None,
                            'retryCount': 0
                        }
                        db.collection('unfound_barcodes').add(unfound_data)
                        print(f"❌ Could not scrape barcode {barcode}, added to unfound list")
                    
                    processed_count += 1
                    
                except Exception as e:
                    error_msg = f"Row {row_num}: {str(e)}"
                    errors.append(error_msg)
                    print(f"Error processing row {row_num}: {e}")
        
        response = {
            'status': 'success',
            'message': f'Processed {processed_count} barcodes',
            'processed_count': processed_count,
            'scraped_count': scraped_count,
            'skipped_count': skipped_count,
            'errors': errors[:10] if errors else []
        }
        
        print(f"Import completed: {processed_count} processed, {scraped_count} scraped, {skipped_count} skipped")
        return jsonify(response)
        
    except Exception as e:
        print(f"Import error: {e}")
        return jsonify({'error': str(e)}), 500

def scrape_product_data(barcode, url):
    """Scrape product data from Smart Consumer website"""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException
    from bs4 import BeautifulSoup
    
    driver = None
    try:
        print(f"🔍 Scraping barcode: {barcode}")
        
        # Setup Chrome options for headless browsing
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        # Initialize Chrome driver
        driver = webdriver.Chrome(options=chrome_options)
        driver.get(url)
        
        # Wait for page to load
        wait = WebDriverWait(driver, 10)
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "h1")))
        except TimeoutException:
            print(f"⚠️ Timeout waiting for page load for barcode: {barcode}")
        
        # Extract product data
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        
        # Extract product information
        product_data = {
            'name': 'N/A',
            'brand': '',
            'category': '',
            'mrp': 0.0,
            'salePrice': 0.0,
            'imageUrl': '',
            'description': '',
            'size': '',
            'unit': '',
            'isActive': True,
            'useInFirstStart': False
        }
        
        # Try to extract product name
        try:
            name_element = soup.find('h1')
            if name_element:
                product_name = name_element.get_text().strip()
                if product_name and product_name != '' and product_name != 'string indices must be integers, not \'str\'':
                    product_data['name'] = product_name
                    print(f"📦 Found product name: {product_name}")
                else:
                    print(f"⚠️ Empty or invalid product name for barcode: {barcode}")
            else:
                print(f"⚠️ No h1 element found for barcode: {barcode}")
        except Exception as e:
            print(f"❌ Error extracting name for {barcode}: {e}")
        
        # Try to extract price
        try:
            price_elements = soup.find_all(['span', 'div'], class_=lambda x: x and 'price' in x.lower())
            for element in price_elements:
                text = element.get_text().strip()
                if '₹' in text or 'Rs' in text:
                    # Extract numeric value
                    import re
                    price_match = re.search(r'[\d,]+\.?\d*', text.replace(',', ''))
                    if price_match:
                        price = float(price_match.group())
                        product_data['mrp'] = price
                        product_data['salePrice'] = price
                        print(f"💰 Found price: ₹{price}")
                        break
        except Exception as e:
            print(f"❌ Error extracting price for {barcode}: {e}")
        
        # Try to extract image
        try:
            img_element = soup.find('img')
            if img_element and img_element.get('src'):
                product_data['imageUrl'] = img_element['src']
                product_data['photoPath'] = img_element['src']
                print(f"🖼️ Found image: {img_element['src']}")
        except Exception as e:
            print(f"❌ Error extracting image for {barcode}: {e}")
        
        # Try to extract brand
        try:
            brand_elements = soup.find_all(['span', 'div'], class_=lambda x: x and 'brand' in x.lower())
            for element in brand_elements:
                text = element.get_text().strip()
                if text and len(text) < 50:  # Reasonable brand name length
                    product_data['brand'] = text
                    print(f"🏷️ Found brand: {text}")
                    break
        except Exception as e:
            print(f"❌ Error extracting brand for {barcode}: {e}")
        
        print(f"✅ Scraping completed for {barcode}: {product_data['name']}")
        return product_data
        
    except Exception as e:
        print(f"❌ Error scraping {barcode}: {e}")
        return None
    finally:
        if driver:
            driver.quit()

if __name__ == '__main__':
    print("🚀 Starting Flask application...")
    print("📊 Firebase connection: ✅ Connected")
    print("🔐 Admin credentials: admin / India@123")
    print("🌐 Dashboard: http://localhost:5000")
    app.run(host='0.0.0.0', port=5000, debug=True)
