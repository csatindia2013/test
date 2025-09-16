from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session, current_app
from flask_cors import CORS
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.exceptions import RequestEntityTooLarge
from datetime import datetime, timezone
import json
import firebase_admin
from firebase_admin import credentials, firestore
import traceback
import openpyxl
from openpyxl import Workbook, load_workbook
from io import BytesIO
import os
import requests
from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
import threading
import schedule
import logging
from logging.handlers import RotatingFileHandler
from dotenv import load_dotenv
from config import config

# Load environment variables
load_dotenv()

# Global variables
db = None
firebase_status = "disconnected"
background_processor = None
processing_status = {
    'running': False,
    'last_run': None,
    'processed_count': 0,
    'success_count': 0,
    'error_count': 0,
    'current_barcode': None
}
processed_barcodes_history = []

# Initialize Firebase globally
def init_firebase_global():
    global db, firebase_status
    try:
        print("Initializing Firebase globally...")
        try:
            existing_app = firebase_admin.get_app()
            print("Firebase app already initialized")
        except ValueError:
            print("Loading service account file...")
            cred = credentials.Certificate('firebase-service-account.json')
            firebase_admin.initialize_app(cred)
            print("Firebase app initialized with service account")
        
        db = firestore.client()
        firebase_status = "connected"
        print("Firebase connection established successfully")
        return True
        
    except Exception as e:
        print(f"Failed to initialize Firebase: {e}")
        import traceback
        traceback.print_exc()
        firebase_status = "disconnected"
        db = None
        return False

# Call the initialization function
init_firebase_global()

# User class for Flask-Login
class User(UserMixin):
    def __init__(self, id, username, role='admin'):
        self.id = id
        self.username = username
        self.role = role

def create_app(config_name=None):
    """Application factory pattern"""
    app = Flask(__name__)
    
    # Load configuration
    config_name = config_name or os.environ.get('FLASK_ENV', 'development')
    app.config.from_object(config[config_name])
    config[config_name].init_app(app)
    
    # Initialize extensions
    CORS(app, origins=app.config['CORS_ORIGINS'])
    
    # Initialize Flask-Login
    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = 'login'
    login_manager.login_message = 'Please log in to access this page.'
    login_manager.login_message_category = 'info'
    
    @login_manager.user_loader
    def load_user(user_id):
        admin_username = app.config['ADMIN_USERNAME']
        admin_role = 'admin'
        user1_username = app.config['USER1_USERNAME']
        user1_role = 'user'
        
        users = {
            admin_username: User(admin_username, admin_username, admin_role),
            user1_username: User(user1_username, user1_username, user1_role)
        }
        return users.get(user_id)
    
    # Initialize rate limiting
    limiter = Limiter(
        app=app,
        key_func=get_remote_address,
        default_limits=[app.config['RATELIMIT_DEFAULT']]
    )
    
    # Initialize Firebase
    # Firebase is now initialized globally, no need to initialize here
    
    # Register blueprints/routes
    register_routes(app, limiter)
    
    # Initialize background processor
    init_background_processor(app)
    
    return app

def init_firebase(app):
    """Initialize Firebase with proper error handling"""
    global db, firebase_status
    
    try:
        print("Initializing Firebase...")
        
        # Simple initialization - just use the service account file
        try:
            existing_app = firebase_admin.get_app()
            print("Firebase app already initialized, using existing app")
        except ValueError:
            print("Loading service account file...")
            cred = credentials.Certificate('firebase-service-account.json')
            firebase_admin.initialize_app(cred)
            print("Firebase app initialized with service account")
        
        db = firestore.client()
        firebase_status = "connected"
        print("Firebase connection established successfully")
        
    except Exception as e:
        print(f"Failed to initialize Firebase: {e}")
        import traceback
        traceback.print_exc()
        firebase_status = "disconnected"
        db = None

def register_routes(app, limiter):
    """Register all application routes"""
    
    # Health check endpoint
    @app.route('/health')
    def health_check():
        """Health check endpoint for load balancers"""
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'firebase_status': firebase_status,
            'version': '1.0.0'
        })
    
    # Test Firebase connection endpoint
    @app.route('/api/test-firebase', methods=['GET'])
    def test_firebase():
        """Test Firebase connection and list collections"""
        try:
            if not db:
                return jsonify({
                    'status': 'error',
                    'message': 'Database not available',
                    'firebase_status': firebase_status
                })
            
            # Try to get collections
            collections = list(db.collections())
            collection_names = [c.id for c in collections]
            
            # Try to get a sample document from barcode_cache
            barcode_cache_ref = db.collection('barcode_cache')
            sample_docs = list(barcode_cache_ref.limit(1).stream())
            
            return jsonify({
                'status': 'success',
                'message': 'Firebase connection working',
                'firebase_status': firebase_status,
                'collections': collection_names,
                'barcode_cache_sample_count': len(sample_docs),
                'sample_doc': sample_docs[0].to_dict() if sample_docs else None
            })
            
        except Exception as e:
            return jsonify({
                'status': 'error',
                'message': f'Firebase test failed: {str(e)}',
                'firebase_status': firebase_status
            })
    
    # Test products endpoint without authentication
    @app.route('/api/test-products', methods=['GET'])
    def test_products():
        """Test products endpoint without authentication"""
        try:
            # Check both collections
            barcode_cache_products = []
            products_collection_products = []
            
            if db:
                # Check barcode_cache collection
                try:
                    barcode_cache_ref = db.collection('barcode_cache')
                    barcode_cache_docs = barcode_cache_ref.stream()
                    for doc in barcode_cache_docs:
                        product_data = doc.to_dict()
                        product_data['id'] = doc.id
                        barcode_cache_products.append(product_data)
                except Exception as e:
                    print(f"Error getting barcode_cache: {e}")
                
                # Check products collection
                try:
                    products_ref = db.collection('products')
                    products_docs = products_ref.stream()
                    for doc in products_docs:
                        product_data = doc.to_dict()
                        product_data['id'] = doc.id
                        products_collection_products.append(product_data)
                except Exception as e:
                    print(f"Error getting products: {e}")
            
            return jsonify({
                'status': 'success',
                'message': 'Products retrieved successfully',
                'firebase_status': firebase_status,
                'barcode_cache_count': len(barcode_cache_products),
                'products_collection_count': len(products_collection_products),
                'barcode_cache_products': barcode_cache_products[:3],  # First 3
                'products_collection_products': products_collection_products[:3]  # First 3
            })
        except Exception as e:
            return jsonify({
                'status': 'error',
                'message': f'Products test failed: {str(e)}',
                'firebase_status': firebase_status
            })
    
    @app.route('/api/test-db-write', methods=['POST'])
    def test_db_write():
        """Test database write operations"""
        try:
            test_data = {
                'barcode': 'TEST123456789',
                'name': 'Test Product',
                'price': '₹100',
                'mrp': '₹120',
                'image': 'https://via.placeholder.com/300x300/cccccc/666666?text=Test',
                'brand': 'Test Brand',
                'category': 'Test Category',
                'description': 'Test Description',
                'verified': False,
                'source': 'test',
                'createdAt': datetime.now().isoformat(),
                'scrapedAt': datetime.now().isoformat()
            }
            
            print(f"DEBUG: Testing database write with data: {test_data}")
            db.collection('barcode_cache').document('TEST123456789').set(test_data)
            print(f"DEBUG: ✅ Successfully wrote test data to barcode_cache")
            
            # Verify it was written
            doc = db.collection('barcode_cache').document('TEST123456789').get()
            if doc.exists:
                print(f"DEBUG: ✅ Verified test data exists in database")
                return jsonify({
                    'status': 'success',
                    'message': 'Database write test successful',
                    'data': doc.to_dict()
                })
            else:
                print(f"DEBUG: ❌ Test data not found in database")
                return jsonify({
                    'status': 'error',
                    'message': 'Test data not found after write'
                })
                
        except Exception as e:
            print(f"DEBUG: ❌ Database write test failed: {e}")
            return jsonify({
                'status': 'error',
                'message': f'Database write test failed: {str(e)}'
            })
    
    # Authentication Routes
    @app.route('/login', methods=['GET', 'POST'])
    @limiter.limit("10 per minute")
    def login():
        if request.method == 'POST':
            data = request.get_json()
            username = data.get('username')
            password = data.get('password')
            
            # Load credentials from config
            admin_username = app.config['ADMIN_USERNAME']
            admin_password_hash = app.config['ADMIN_PASSWORD_HASH']
            user1_username = app.config['USER1_USERNAME']
            user1_password_hash = app.config['USER1_PASSWORD_HASH']
            
            users = {
                admin_username: admin_password_hash,
                user1_username: user1_password_hash
            }
            
            if username in users and check_password_hash(users[username], password):
                user = User(username, username)
                login_user(user)
                app.logger.info(f"User {username} logged in successfully")
                return jsonify({
                    'status': 'success',
                    'message': 'Login successful',
                    'user': {
                        'id': user.id,
                        'username': user.username,
                        'role': user.role
                    }
                })
            else:
                app.logger.warning(f"Failed login attempt for username: {username}")
                return jsonify({
                    'status': 'error',
                    'message': 'Invalid username or password'
                }), 401
        
        return render_template('login.html')

    @app.route('/logout', methods=['POST'])
    @login_required
    def logout():
        app.logger.info(f"User {current_user.username} logged out")
        logout_user()
        return jsonify({
            'status': 'success',
            'message': 'Logged out successfully'
        })

    @app.route('/api/auth/status', methods=['GET'])
    def auth_status():
        if current_user.is_authenticated:
            return jsonify({
                'status': 'success',
                'authenticated': True,
                'user': {
                    'id': current_user.id,
                    'username': current_user.username,
                    'role': current_user.role
                }
            })
        else:
            return jsonify({
                'status': 'success',
                'authenticated': False
            })

    @app.route('/')
    @login_required
    def index():
        return render_template('index.html')
    
    # Error handlers
    @app.errorhandler(400)
    def bad_request(e):
        app.logger.warning(f"Bad request: {e}")
        return jsonify({'error': 'Bad request'}), 400
    
    @app.errorhandler(401)
    def unauthorized(e):
        app.logger.warning(f"Unauthorized access attempt: {e}")
        return jsonify({'error': 'Unauthorized'}), 401
    
    @app.errorhandler(403)
    def forbidden(e):
        app.logger.warning(f"Forbidden access attempt: {e}")
        return jsonify({'error': 'Forbidden'}), 403
    
    @app.errorhandler(404)
    def not_found(e):
        app.logger.info(f"Not found: {e}")
        return jsonify({'error': 'Not found'}), 404
    
    @app.errorhandler(413)
    def too_large(e):
        app.logger.warning(f"File too large: {e}")
        return jsonify({'error': 'File too large'}), 413
    
    @app.errorhandler(429)
    def ratelimit_handler(e):
        app.logger.warning(f"Rate limit exceeded: {e}")
        return jsonify({'error': 'Rate limit exceeded'}), 429
    
    @app.errorhandler(500)
    def internal_error(e):
        app.logger.error(f"Internal server error: {e}")
        return jsonify({'error': 'Internal server error'}), 500
    
    @app.errorhandler(502)
    def bad_gateway(e):
        app.logger.error(f"Bad gateway: {e}")
        return jsonify({'error': 'Service temporarily unavailable'}), 502
    
    @app.errorhandler(503)
    def service_unavailable(e):
        app.logger.error(f"Service unavailable: {e}")
        return jsonify({'error': 'Service temporarily unavailable'}), 503

def init_background_processor(app):
    """Initialize background processor if enabled"""
    if app.config['BACKGROUND_PROCESSOR_ENABLED']:
        # Background processor initialization code here
        pass

# Create the app instance
app = create_app()

# Background Processing System
background_processor = None
processing_status = {
    'running': False,
    'last_run': None,
    'processed_count': 0,
    'success_count': 0,
    'error_count': 0,
    'current_barcode': None
}

# Processed barcodes history (in-memory storage)
processed_barcodes_history = []

# Mock data for testing when Firebase is not available
MOCK_PRODUCTS = [
    {
        'id': 'mock_1',
        'name': 'Sample Product 1',
        'category': 'Electronics',
        'mrp': 100.0,
        'price': 90.0,
        'useInFirstStart': True,
        'imageUrl': 'https://via.placeholder.com/150',
        'stockQuantity': 10
    },
    {
        'id': 'mock_2',
        'name': 'Sample Product 2',
        'category': 'Clothing',
        'mrp': 50.0,
        'price': 45.0,
        'useInFirstStart': False,
        'imageUrl': 'https://via.placeholder.com/150',
        'stockQuantity': 5
    }
]

MOCK_CATEGORIES = [
    {
        'id': 'mock_cat_1',
        'name': 'Electronics',
        'description': 'Electronic devices and accessories',
        'isActive': True
    },
    {
        'id': 'mock_cat_2',
        'name': 'Clothing',
        'description': 'Apparel and fashion items',
        'isActive': True
    }
]

MOCK_UNFOUND_BARCODES = [
    {
        'id': 'mock_barcode_1',
        'barcode': '1234567890123',
        'timestamp': '2025-01-13T10:30:00Z',
        'deviceId': 'device_001',
        'location': 'Store A'
    },
    {
        'id': 'mock_barcode_2',
        'barcode': '9876543210987',
        'timestamp': '2025-01-13T11:15:00Z',
        'deviceId': 'device_002',
        'location': 'Store B'
    }
]

try:
    print("Initializing Firebase...")
    
    # Try to load service account file or environment variables
    try:
        # Check if Firebase is already initialized
        try:
            existing_app = firebase_admin.get_app()
            print("Firebase app already initialized, using existing app")
        except ValueError:
            print("Loading service account file...")
            cred = credentials.Certificate('firebase-service-account.json')
            firebase_admin.initialize_app(cred)
            print("Firebase app initialized with service account")
    except Exception as e:
        print(f"Service account file failed: {e}")
        
        # Try environment variables (for Render deployment)
        try:
            import os
            # Check if Firebase is already initialized
            try:
                existing_app = firebase_admin.get_app()
                print("Firebase app already initialized, using existing app")
            except ValueError:
                firebase_config = {
                "type": "service_account",
                "project_id": os.environ.get('FIREBASE_PROJECT_ID'),
                "private_key_id": os.environ.get('FIREBASE_PRIVATE_KEY_ID'),
                "private_key": os.environ.get('FIREBASE_PRIVATE_KEY', '').replace('\\n', '\n'),
                "client_email": os.environ.get('FIREBASE_CLIENT_EMAIL'),
                "client_id": os.environ.get('FIREBASE_CLIENT_ID'),
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
                "client_x509_cert_url": f"https://www.googleapis.com/robot/v1/metadata/x509/{os.environ.get('FIREBASE_CLIENT_EMAIL')}"
            }
            
            if all(firebase_config.values()):
                cred = credentials.Certificate(firebase_config)
                firebase_admin.initialize_app(cred)
                print("Firebase app initialized with environment variables")
            else:
                raise Exception("Missing Firebase environment variables")
                
        except Exception as e2:
            print(f"Environment variables failed: {e2}")
            print("Trying Application Default Credentials...")
            try:
                existing_app = firebase_admin.get_app()
                print("Firebase app already initialized, using existing app")
            except ValueError:
                firebase_admin.initialize_app()
                print("Firebase app initialized with default credentials")
    
    # Create Firestore client
    db = firestore.client()
    print("Firebase Firestore client created successfully")
    
    # Test connection with timeout
    print("Testing Firebase connection...")
    try:
        # Simple test - just try to get collections (read-only operation)
        collections = list(db.collections())
        print(f"Firebase connection test successful! Found {len(collections)} collections")
        firebase_status = "connected"
    except Exception as e:
        print(f"Firebase connection test failed: {e}")
        firebase_status = "test_failed"
        db = None
        
except Exception as e:
    print(f"Firebase initialization failed: {e}")
    print(f"Error type: {type(e).__name__}")
    firebase_status = "initialization_failed"
    db = None

# Product Service
class ProductService:
    @staticmethod
    def get_products():
        print(f"ProductService.get_products() called - Firebase status: {firebase_status}")
        print(f"Database object: {db}")
        
        if not db:
            print("Database not available, returning error")
            return {"error": "Database not available", "status": firebase_status}
        
        try:
            print("Getting products from Firebase barcode_cache collection...")
            products_ref = db.collection('barcode_cache')
            docs = products_ref.stream()
            
            products = []
            doc_count = 0
            for doc in docs:
                doc_count += 1
                product_data = doc.to_dict()
                product_data['id'] = doc.id
                
                print(f"Processing document {doc_count}: {doc.id}")
                
                # Map Firebase fields to expected dashboard fields
                mapped_product = {
                    'id': product_data.get('id', doc.id),
                    'name': product_data.get('name', 'Unnamed Product'),
                    'brand': product_data.get('brand', ''),
                    'category': product_data.get('category', ''),
                    'barcode': doc.id,  # In barcode_cache, the document ID is the barcode
                    'mrp': product_data.get('mrp', 0),
                    'salePrice': product_data.get('salePrice', 0),
                    'stock': product_data.get('stockQuantity', 0),
                    'isActive': product_data.get('isActive', True),
                    'useInFirstStart': product_data.get('useInFirstStart', False),
                    'imageUrl': product_data.get('photoPath', ''),  # barcode_cache uses photoPath
                    'description': product_data.get('description', ''),
                    'createdAt': product_data.get('createdAt', ''),
                    'updatedAt': product_data.get('updatedAt', ''),
                    # Additional fields from barcode_cache
                    'size': product_data.get('size', ''),
                    'unit': product_data.get('unit', ''),
                    'scanCount': product_data.get('scanCount', 0),
                    'syncStatus': product_data.get('syncStatus', ''),
                    'sortOrder': product_data.get('sortOrder', 0),
                    # Keep original fields for reference
                    'originalData': product_data
                }
                products.append(mapped_product)
            
            print(f"Retrieved {len(products)} products from Firebase barcode_cache collection")
            if len(products) == 0:
                print("No products found in barcode_cache collection")
                # Try to check if the collection exists by attempting to get collection info
                try:
                    collections = db.collections()
                    print(f"Available collections: {[c.id for c in collections]}")
                except Exception as e:
                    print(f"Error getting collections: {e}")
            
            return products
        except Exception as e:
            print(f"Error getting products: {e}")
            import traceback
            traceback.print_exc()
            return {"error": str(e), "status": "error"}

    @staticmethod
    def get_product(product_id):
        if not db:
            return {"error": "Database not available", "status": firebase_status}
        
        try:
            doc_ref = db.collection('products').document(product_id)
            doc = doc_ref.get()
            
            if doc.exists:
                product_data = doc.to_dict()
                product_data['id'] = doc.id
                return product_data
            else:
                return {"error": "Product not found"}
        except Exception as e:
            return {"error": str(e)}

    @staticmethod
    def create_product(product_data):
        if not db:
            return {"error": "Database not available", "status": firebase_status}
        
        try:
            product_data['createdAt'] = datetime.now().isoformat()
            product_data['updatedAt'] = datetime.now().isoformat()
            
            # Use barcode as document ID if available, otherwise generate random ID
            if 'barcode' in product_data and product_data['barcode']:
                doc_ref = db.collection('products').document(product_data['barcode']).set(product_data)
                return {"id": product_data['barcode'], "message": "Product created successfully"}
            else:
                doc_ref = db.collection('products').add(product_data)
                return {"id": doc_ref[1].id, "message": "Product created successfully"}
        except Exception as e:
            return {"error": str(e)}

    @staticmethod
    def update_product(product_id, product_data):
        if not db:
            return {"error": "Database not available", "status": firebase_status}
        
        try:
            product_data['updatedAt'] = datetime.now().isoformat()
            
            doc_ref = db.collection('products').document(product_id)
            doc_ref.update(product_data)
            return {"message": "Product updated successfully"}
        except Exception as e:
            return {"error": str(e)}

    @staticmethod
    def delete_product(product_id):
        if not db:
            return {"error": "Database not available", "status": firebase_status}
        
        try:
            # Delete from barcode_cache collection (where main app reads from)
            db.collection('barcode_cache').document(product_id).delete()
            
            # Also try to delete from products collection if it exists there
            try:
                db.collection('products').document(product_id).delete()
            except:
                pass  # Ignore if not found in products collection
            
            return {"message": "Product deleted successfully"}
        except Exception as e:
            return {"error": str(e)}

# Category Service
class CategoryService:
    @staticmethod
    def get_categories():
        if not db:
            return {"error": "Database not available", "status": firebase_status}
        
        try:
            print("Getting categories from Firebase...")
            categories_ref = db.collection('categories')
            docs = categories_ref.stream()
            
            categories = []
            for doc in docs:
                category_data = doc.to_dict()
                category_data['id'] = doc.id
                
                # Map Firebase fields to expected dashboard fields
                mapped_category = {
                    'id': category_data.get('id', doc.id),
                    'name': category_data.get('name', 'Unnamed Category'),
                    'description': category_data.get('description', ''),
                    'isActive': category_data.get('isActive', True),
                    'productCount': category_data.get('productCount', 0),
                    'createdAt': category_data.get('createdAt', ''),
                    'updatedAt': category_data.get('updatedAt', ''),
                    # Keep original fields for reference
                    'originalData': category_data
                }
                categories.append(mapped_category)
            
            print(f"Retrieved {len(categories)} categories from Firebase")
            return categories
        except Exception as e:
            print(f"Error getting categories: {e}")
            return {"error": str(e), "status": "error"}

    @staticmethod
    def create_category(category_data):
        if not db:
            return {"error": "Database not available", "status": firebase_status}
        
        try:
            category_data['createdAt'] = datetime.now().isoformat()
            category_data['updatedAt'] = datetime.now().isoformat()
            
            doc_ref = db.collection('categories').add(category_data)
            return {"id": doc_ref[1].id, "message": "Category created successfully"}
        except Exception as e:
            return {"error": str(e)}

    @staticmethod
    def update_category(category_id, category_data):
        if not db:
            return {"error": "Database not available", "status": firebase_status}
        
        try:
            category_data['updatedAt'] = datetime.now().isoformat()
            
            doc_ref = db.collection('categories').document(category_id)
            doc_ref.update(category_data)
            return {"message": "Category updated successfully"}
        except Exception as e:
            return {"error": str(e)}

    @staticmethod
    def delete_category(category_id):
        if not db:
            return {"error": "Database not available", "status": firebase_status}
        
        try:
            db.collection('categories').document(category_id).delete()
            return {"message": "Category deleted successfully"}
        except Exception as e:
            return {"error": str(e)}

# Routes are now defined in the register_routes function

# Product Routes
@app.route('/api/products', methods=['GET'])
@login_required
def get_products():
    products = ProductService.get_products()
    return jsonify(products)

@app.route('/api/products/<product_id>', methods=['GET'])
def get_product(product_id):
    product = ProductService.get_product(product_id)
    return jsonify(product)

@app.route('/api/products', methods=['POST'])
def create_product():
    product_data = request.get_json()
    result = ProductService.create_product(product_data)
    return jsonify(result)

@app.route('/api/products/<product_id>', methods=['PUT'])
def update_product(product_id):
    product_data = request.get_json()
    result = ProductService.update_product(product_id, product_data)
    return jsonify(result)

@app.route('/api/products/bulk-delete', methods=['DELETE'])
def bulk_delete_products():
    try:
        data = request.get_json()
        if not data or 'product_ids' not in data:
            return jsonify({'error': 'Product IDs required'}), 400
        
        product_ids = data['product_ids']
        if not isinstance(product_ids, list) or len(product_ids) == 0:
            return jsonify({'error': 'Product IDs must be a non-empty list'}), 400
        
        if db:
            # Delete products from Firebase (both barcode_cache and products collections)
            deleted_count = 0
            for product_id in product_ids:
                try:
                    # Delete from barcode_cache collection (where main app reads from)
                    db.collection('barcode_cache').document(product_id).delete()
                    
                    # Also try to delete from products collection if it exists there
                    try:
                        db.collection('products').document(product_id).delete()
                    except:
                        pass  # Ignore if not found in products collection
                    
                    deleted_count += 1
                except Exception as e:
                    print(f"Error deleting product {product_id}: {e}")
                    continue
        
            return jsonify({
                'message': f'Successfully deleted {deleted_count} out of {len(product_ids)} products',
                'deleted_count': deleted_count,
                'total_requested': len(product_ids)
            })
        else:
            return jsonify({'error': 'Database not available'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories/bulk-delete', methods=['DELETE'])
def bulk_delete_categories():
    try:
        data = request.get_json()
        if not data or 'category_ids' not in data:
            return jsonify({'error': 'Category IDs required'}), 400
        
        category_ids = data['category_ids']
        if not isinstance(category_ids, list) or len(category_ids) == 0:
            return jsonify({'error': 'Category IDs must be a non-empty list'}), 400
        
        if db:
            # Delete categories from Firebase
            deleted_count = 0
            for category_id in category_ids:
                try:
                    db.collection('categories').document(category_id).delete()
                    deleted_count += 1
                except Exception as e:
                    print(f"Error deleting category {category_id}: {e}")
                    continue
        
            return jsonify({
                'message': f'Successfully deleted {deleted_count} out of {len(category_ids)} categories',
                'deleted_count': deleted_count,
                'total_requested': len(category_ids)
            })
        else:
            return jsonify({'error': 'Database not available'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Excel Export Endpoints
@app.route('/api/products/export', methods=['GET'])
def export_products():
    try:
        if db:
            # Get products from Firebase
            products_ref = db.collection('products')
            products = []
            for doc in products_ref.stream():
                product_data = doc.to_dict()
                product_data['id'] = doc.id
                products.append(product_data)
        else:
            # Use mock data if Firebase not available
            products = MOCK_PRODUCTS
        
        # Create Excel file in memory using openpyxl
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Define column order
        column_order = ['id', 'name', 'category', 'mrp', 'price', 'useInFirstStart', 'imageUrl', 'stockQuantity']
        
        # Add headers
        for col, header in enumerate(column_order, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, product in enumerate(products, 2):
            for col, header in enumerate(column_order, 1):
                ws.cell(row=row, column=col, value=product.get(header, ''))
        
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"products_export_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories/export', methods=['GET'])
def export_categories():
    try:
        if db:
            # Get categories from Firebase
            categories_ref = db.collection('categories')
            categories = []
            for doc in categories_ref.stream():
                category_data = doc.to_dict()
                category_data['id'] = doc.id
                categories.append(category_data)
        else:
            # Use mock data if Firebase not available
            categories = MOCK_CATEGORIES
        
        # Create Excel file in memory using openpyxl
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Categories"
        
        # Define column order
        column_order = ['id', 'name', 'description', 'isActive']
        
        # Add headers
        for col, header in enumerate(column_order, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, category in enumerate(categories, 2):
            for col, header in enumerate(column_order, 1):
                ws.cell(row=row, column=col, value=category.get(header, ''))
        
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"categories_export_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Excel Import Endpoints
@app.route('/api/products/import', methods=['POST'])
def import_products():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'File must be an Excel file (.xlsx or .xls)'}), 400
        
        # Read Excel file using openpyxl
        wb = load_workbook(file)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Validate required columns
        required_columns = ['name', 'category', 'mrp', 'price']
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            return jsonify({'error': f'Missing required columns: {", ".join(missing_columns)}'}), 400
        
        # Process and import products
        imported_count = 0
        errors = []
        
        for row_num in range(2, ws.max_row + 1):
            try:
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[header] = cell_value
                
                # Helper function to check if value is not None/empty
                def is_not_empty(value):
                    return value is not None and str(value).strip() != ''
                
                product_data = {
                    'name': str(row_data['name']),
                    'category': str(row_data['category']),
                    'mrp': float(row_data['mrp']) if is_not_empty(row_data['mrp']) else 0.0,
                    'price': float(row_data['price']) if is_not_empty(row_data['price']) else 0.0,
                    'useInFirstStart': bool(row_data.get('useInFirstStart', False)) if is_not_empty(row_data.get('useInFirstStart')) else False,
                    'imageUrl': str(row_data.get('imageUrl', '')) if is_not_empty(row_data.get('imageUrl')) else '',
                    'stockQuantity': int(row_data.get('stockQuantity', 0)) if is_not_empty(row_data.get('stockQuantity')) else 0
                }
                
                if db:
                    # Add to Firebase
                    db.collection('products').add(product_data)
                else:
                    # Add to mock data (for testing)
                    product_data['id'] = f"mock_{len(MOCK_PRODUCTS) + 1}"
                    MOCK_PRODUCTS.append(product_data)
            
                imported_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        response = {
            'message': f'Successfully imported {imported_count} products',
            'imported_count': imported_count,
            'total_rows': ws.max_row - 1
        }
        
        if errors:
            response['errors'] = errors[:10]  # Limit to first 10 errors
            if len(errors) > 10:
                response['errors'].append(f"... and {len(errors) - 10} more errors")
        
        return jsonify(response)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories/import', methods=['POST'])
def import_categories():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'File must be an Excel file (.xlsx or .xls)'}), 400
        
        # Read Excel file using openpyxl
        wb = load_workbook(file)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Validate required columns
        required_columns = ['name']
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            return jsonify({'error': f'Missing required columns: {", ".join(missing_columns)}'}), 400
        
        # Process and import categories
        imported_count = 0
        errors = []
        
        for row_num in range(2, ws.max_row + 1):
            try:
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[header] = cell_value
                
                # Helper function to check if value is not None/empty
                def is_not_empty(value):
                    return value is not None and str(value).strip() != ''
                
                category_data = {
                    'name': str(row_data['name']),
                    'description': str(row_data.get('description', '')) if is_not_empty(row_data.get('description')) else '',
                    'isActive': bool(row_data.get('isActive', True)) if is_not_empty(row_data.get('isActive')) else True
                }
                
                if db:
                    # Add to Firebase
                    db.collection('categories').add(category_data)
                else:
                    # Add to mock data (for testing)
                    category_data['id'] = f"mock_cat_{len(MOCK_CATEGORIES) + 1}"
                    MOCK_CATEGORIES.append(category_data)
                
                imported_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        response = {
            'message': f'Successfully imported {imported_count} categories',
            'imported_count': imported_count,
            'total_rows': ws.max_row - 1
        }
        
        if errors:
            response['errors'] = errors[:10]  # Limit to first 10 errors
            if len(errors) > 10:
                response['errors'].append(f"... and {len(errors) - 10} more errors")
        
        return jsonify(response)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Unfound Barcode Management Endpoints
@app.route('/api/unfound-barcodes', methods=['GET'])
def get_unfound_barcodes():
    try:
        if db:
            # Get unfound barcodes from Firebase
            unfound_barcodes_ref = db.collection('unfound_barcodes')
            unfound_barcodes = []
            
            for doc in unfound_barcodes_ref.stream():
                barcode_data = doc.to_dict()
                barcode_data['id'] = doc.id
                
                # Determine source based on available fields
                source = barcode_data.get('source', 'Unknown')
                if source == 'Unknown':
                    # Check if it has Excel-specific fields
                    if 'deviceId' in barcode_data and barcode_data.get('deviceId') != 'Unknown':
                        source = 'Firebase DB'
                    elif 'location' in barcode_data and barcode_data.get('location') != 'Unknown':
                        source = 'Firebase DB'
                    else:
                        source = 'Excel Import'
                elif source == 'excel':
                    source = 'Excel Import'
                elif source == 'firebase_db' or source == 'firebase':
                    source = 'Firebase DB'
                
                barcode_data['source'] = source
                unfound_barcodes.append(barcode_data)
            
            # Sort by source and then by creation date
            unfound_barcodes.sort(key=lambda x: (x.get('source', ''), x.get('createdAt', '')), reverse=True)
            
            # Group by source for better organization
            grouped_barcodes = {
                'excel': [],
                'firebase_db': [],
                'total_count': len(unfound_barcodes)
            }
            
            for barcode in unfound_barcodes:
                if barcode['source'] == 'Excel Import':
                    grouped_barcodes['excel'].append(barcode)
                elif barcode['source'] == 'Firebase DB':
                    grouped_barcodes['firebase_db'].append(barcode)
            
            # Return the array directly for backward compatibility with existing frontend
            return jsonify(unfound_barcodes)
        else:
            # Use mock data if Firebase not available
            mock_barcodes = []
            for i, barcode in enumerate(MOCK_UNFOUND_BARCODES):
                barcode_copy = barcode.copy()
                barcode_copy['source'] = 'Excel Import' if i % 2 == 0 else 'Firebase DB'
                mock_barcodes.append(barcode_copy)
            
            # Return the array directly for backward compatibility with existing frontend
            return jsonify(mock_barcodes)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/unfound-barcodes/sources', methods=['GET'])
def get_unfound_barcodes_sources():
    """Get unfound barcodes grouped by source with counts"""
    try:
        if db:
            # Get unfound barcodes from Firebase
            unfound_barcodes_ref = db.collection('unfound_barcodes')
            unfound_barcodes = []
            
            for doc in unfound_barcodes_ref.stream():
                barcode_data = doc.to_dict()
                barcode_data['id'] = doc.id
                
                # Determine source based on available fields
                source = barcode_data.get('source', 'Unknown')
                if source == 'Unknown':
                    # Check if it has Excel-specific fields
                    if 'deviceId' in barcode_data and barcode_data.get('deviceId') != 'Unknown':
                        source = 'Firebase DB'
                    elif 'location' in barcode_data and barcode_data.get('location') != 'Unknown':
                        source = 'Firebase DB'
                    else:
                        source = 'Excel Import'
                elif source == 'excel':
                    source = 'Excel Import'
                elif source == 'firebase_db' or source == 'firebase':
                    source = 'Firebase DB'
                
                barcode_data['source'] = source
                unfound_barcodes.append(barcode_data)
            
            # Group by source
            excel_barcodes = [b for b in unfound_barcodes if b['source'] == 'Excel Import']
            firebase_barcodes = [b for b in unfound_barcodes if b['source'] == 'Firebase DB']
            
            return jsonify({
                'sources': {
                    'excel': {
                        'name': 'Excel Import',
                        'count': len(excel_barcodes),
                        'barcodes': excel_barcodes
                    },
                    'firebase_db': {
                        'name': 'Firebase DB',
                        'count': len(firebase_barcodes),
                        'barcodes': firebase_barcodes
                    }
                },
                'total_count': len(unfound_barcodes),
                'summary': {
                    'excel_count': len(excel_barcodes),
                    'firebase_db_count': len(firebase_barcodes),
                    'total': len(unfound_barcodes)
                }
            })
        else:
            return jsonify({'error': 'Database not available'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/unfound-barcodes', methods=['POST'])
def create_unfound_barcode():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Validate required fields
        if 'barcode' not in data:
            return jsonify({'error': 'Barcode is required'}), 400
        
        barcode_data = {
            'barcode': data['barcode'],
            'source': data.get('source', 'firebase_db'),  # Default to 'firebase_db' for manual additions
            'createdAt': data.get('timestamp', datetime.now().isoformat()),
            'timestamp': data.get('timestamp', datetime.now().isoformat()),  # Keep for backward compatibility
            'deviceId': data.get('deviceId', 'Unknown'),
            'location': data.get('location', 'Unknown'),
            'status': data.get('status', 'pending'),
            'lastRetry': None,
            'retryCount': 0
        }
        
        if db:
            # Add to Firebase
            doc_ref = db.collection('unfound_barcodes').add(barcode_data)
            barcode_data['id'] = doc_ref[1].id
        else:
            # Add to mock data
            barcode_data['id'] = f"mock_barcode_{len(MOCK_UNFOUND_BARCODES) + 1}"
            MOCK_UNFOUND_BARCODES.append(barcode_data)
            
        return jsonify({
            'message': 'Unfound barcode created successfully',
            'barcode': barcode_data
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/unfound-barcodes/<barcode_id>', methods=['DELETE'])
def delete_unfound_barcode(barcode_id):
    try:
        if db:
            # Delete from Firebase
            db.collection('unfound_barcodes').document(barcode_id).delete()
            return jsonify({'message': 'Unfound barcode deleted successfully'})
        else:
            # Remove from mock data
            global MOCK_UNFOUND_BARCODES
            MOCK_UNFOUND_BARCODES = [barcode for barcode in MOCK_UNFOUND_BARCODES if barcode['id'] != barcode_id]
            return jsonify({'message': 'Unfound barcode deleted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/unfound-barcodes/bulk-delete', methods=['DELETE'])
def bulk_delete_unfound_barcodes():
    try:
        data = request.get_json()
        if not data or 'barcode_ids' not in data:
            return jsonify({'error': 'Barcode IDs required'}), 400
        
        barcode_ids = data['barcode_ids']
        if not isinstance(barcode_ids, list) or len(barcode_ids) == 0:
            return jsonify({'error': 'Barcode IDs must be a non-empty list'}), 400
        
        if db:
            # Delete from Firebase
            deleted_count = 0
            for barcode_id in barcode_ids:
                try:
                    db.collection('unfound_barcodes').document(barcode_id).delete()
                    deleted_count += 1
                except Exception as e:
                    print(f"Error deleting unfound barcode {barcode_id}: {e}")
                    continue
        
            return jsonify({
                'message': f'Successfully deleted {deleted_count} out of {len(barcode_ids)} unfound barcodes',
                'deleted_count': deleted_count,
                'total_requested': len(barcode_ids)
            })
        else:
            # Remove from mock data
            global MOCK_UNFOUND_BARCODES
            original_count = len(MOCK_UNFOUND_BARCODES)
            MOCK_UNFOUND_BARCODES = [barcode for barcode in MOCK_UNFOUND_BARCODES if barcode['id'] not in barcode_ids]
            deleted_count = original_count - len(MOCK_UNFOUND_BARCODES)
        
            return jsonify({
                'message': f'Successfully deleted {deleted_count} out of {len(barcode_ids)} unfound barcodes',
                'deleted_count': deleted_count,
                'total_requested': len(barcode_ids)
            })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/unfound-barcodes/export', methods=['GET'])
def export_unfound_barcodes():
    try:
        if db:
            # Get unfound barcodes from Firebase
            unfound_barcodes_ref = db.collection('unfound_barcodes')
            unfound_barcodes = []
            for doc in unfound_barcodes_ref.stream():
                barcode_data = doc.to_dict()
                barcode_data['id'] = doc.id
                unfound_barcodes.append(barcode_data)
        else:
            # Use mock data if Firebase not available
            unfound_barcodes = MOCK_UNFOUND_BARCODES
        
        # Create Excel file in memory using openpyxl
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Unfound Barcodes"
        
        # Define column order
        column_order = ['id', 'barcode', 'timestamp', 'deviceId', 'location']
        
        # Add headers
        for col, header in enumerate(column_order, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, barcode_data in enumerate(unfound_barcodes, 2):
            for col, header in enumerate(column_order, 1):
                ws.cell(row=row, column=col, value=barcode_data.get(header, ''))
        
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"unfound_barcodes_export_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/products/<product_id>/field', methods=['PATCH'])
def update_product_field(product_id):
    try:
        data = request.get_json()
        if not data or 'field' not in data or 'value' not in data:
            return jsonify({'error': 'Field and value required'}), 400
        
        field = data['field']
        value = data['value']
        
        if db:
            # Update specific field in Firebase
            product_ref = db.collection('products').document(product_id)
            product_ref.update({field: value})
            return jsonify({'message': f'{field} updated successfully'})
        else:
            return jsonify({'error': 'Database not available'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/products/<product_id>', methods=['DELETE'])
def delete_product(product_id):
    result = ProductService.delete_product(product_id)
    return jsonify(result)

# Category Routes
@app.route('/api/categories', methods=['GET'])
def get_categories():
    categories = CategoryService.get_categories()
    return jsonify(categories)

@app.route('/api/categories', methods=['POST'])
def create_category():
    category_data = request.get_json()
    result = CategoryService.create_category(category_data)
    return jsonify(result)

@app.route('/api/categories/<category_id>', methods=['PUT'])
def update_category(category_id):
    category_data = request.get_json()
    result = CategoryService.update_category(category_id, category_data)
    return jsonify(result)

@app.route('/api/categories/<category_id>', methods=['DELETE'])
def delete_category(category_id):
    result = CategoryService.delete_category(category_id)
    return jsonify(result)

# Debug Routes
@app.route('/api/debug/firebase', methods=['GET'])
def debug_firebase():
        return jsonify({
        "status": firebase_status,
        "database_available": db is not None,
        "timestamp": datetime.now().isoformat(),
        "message": f"Firebase status: {firebase_status}"
    })

# Background Processor API Endpoints
@app.route('/api/background-processor/status', methods=['GET'])
def get_background_processor_status():
    """Get current status of background processor"""
    return jsonify({
        'status': 'success',
        'data': processing_status
    })

@app.route('/api/background-processor/start', methods=['POST'])
def start_background_processor_api():
    """Start the background processor"""
    try:
        success = start_background_processor()
        if success:
            return jsonify({
                'status': 'success',
                'message': 'Background processor started successfully'
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'Background processor is already running'
            })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to start background processor: {str(e)}'
        })

@app.route('/api/background-processor/stop', methods=['POST'])
def stop_background_processor_api():
    """Stop the background processor"""
    try:
        stop_background_processor()
        return jsonify({
            'status': 'success',
            'message': 'Background processor stopped successfully'
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to stop background processor: {str(e)}'
        })

@app.route('/api/background-processor/start-continuous', methods=['POST'])
def start_continuous_background_processor():
    """Start the background processor in continuous real-time mode"""
    try:
        success = start_background_processor()
        if success:
            return jsonify({
                'status': 'success',
                'message': 'Background processor started in CONTINUOUS REAL-TIME mode',
                'mode': 'continuous'
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'Background processor is already running'
            })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to start continuous background processor: {str(e)}'
        })

@app.route('/api/background-processor/run-now', methods=['POST'])
def run_background_processor_now():
    """Run background processor immediately (for testing)"""
    try:
        if not processing_status['running']:
            return jsonify({
                'status': 'error',
                'message': 'Background processor is not running. Start it first with /api/background-processor/start-continuous'
            })
        
        # Trigger immediate processing
        unfound_barcodes = get_unfound_barcodes_for_processing()
        
        if unfound_barcodes:
            return jsonify({
                'status': 'success',
                'message': f'Found {len(unfound_barcodes)} barcodes to process. Processing will continue automatically.',
                'barcodes_found': len(unfound_barcodes)
            })
        else:
            return jsonify({
                'status': 'success',
                'message': 'No unfound barcodes to process at this time',
                'barcodes_found': 0
            })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to run background processor: {str(e)}'
        })

@app.route('/api/background-processor/processed-barcodes', methods=['GET'])
def get_processed_barcodes():
    """Get processed barcodes history"""
    try:
        return jsonify({
            'status': 'success',
            'data': processed_barcodes_history
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to get processed barcodes: {str(e)}'
        })

@app.route('/api/background-processor/clear-processed-history', methods=['POST'])
def clear_processed_history():
    """Clear processed barcodes history"""
    try:
        global processed_barcodes_history
        processed_barcodes_history = []
        return jsonify({
            'status': 'success',
            'message': 'Processed barcodes history cleared successfully'
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to clear processed history: {str(e)}'
        })

@app.route('/api/test/create-sample-product', methods=['POST'])
def create_sample_product():
    if not db:
        return jsonify({"error": "Database not available", "status": firebase_status})
    
    try:
        sample_product = {
            'name': 'Test Product from Dashboard',
            'description': 'This is a test product created from the dashboard',
            'price': 99.99,
            'stock': 10,
            'category': 'Test Category',
            'brand': 'Test Brand',
            'barcode': 'TEST123456',
            'isActive': True,
            'createdAt': datetime.now().isoformat(),
            'updatedAt': datetime.now().isoformat()
        }
        
        doc_ref = db.collection('products').document(sample_product['barcode']).set(sample_product)
        return jsonify({
            "status": "success",
            "message": "Sample product created successfully",
            "product_id": sample_product['barcode']
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        })

# Recently Added Products API Endpoints
@app.route('/api/recently-added-products', methods=['GET'])
def get_recently_added_products():
    """Get recently added products from background processor"""
    try:
        if not db:
            return jsonify({
                'status': 'error',
                'message': 'Database not available'
            }), 500
        
        # Get recently added products
        recently_added_ref = db.collection('recently_added_products')
        recently_added_products = []
        
        for doc in recently_added_ref.stream():
            product_data = doc.to_dict()
            product_data['id'] = doc.id
            recently_added_products.append(product_data)
        
        # Sort by addedAt (newest first)
        recently_added_products.sort(key=lambda x: x.get('addedAt', ''), reverse=True)
        
        return jsonify({
            'status': 'success',
            'data': recently_added_products
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to get recently added products: {str(e)}'
        }), 500

@app.route('/api/migrate-products-to-barcode-cache', methods=['POST'])
def migrate_products_to_barcode_cache():
    """Migrate all products from products collection to barcode_cache collection"""
    try:
        if not db:
            return jsonify({'status': 'error', 'message': 'Database not available'}), 500
        
        migrated_count = 0
        
        # Get all products from products collection
        products_ref = db.collection('products')
        products_docs = products_ref.stream()
        
        for doc in products_docs:
            product_data = doc.to_dict()
            barcode = product_data.get('barcode')
            
            if barcode:
                # Check if product already exists in barcode_cache
                existing_product = db.collection('barcode_cache').document(barcode).get()
                
                if not existing_product.exists:
                    # Create barcode_cache data
                    barcode_cache_data = {
                        'barcode': barcode,
                        'name': product_data.get('name', 'Unknown'),
                        'price': product_data.get('price', 'N/A'),
                        'mrp': product_data.get('mrp', product_data.get('price', 'N/A')),
                        'image': product_data.get('image'),
                        'brand': product_data.get('brand', ''),
                        'category': product_data.get('category', ''),
                        'description': product_data.get('description', ''),
                        'source': product_data.get('source', 'migrated_from_products'),
                        'verified': product_data.get('verified', True),
                        'verifiedAt': product_data.get('verifiedAt', datetime.now().isoformat()),
                        'createdAt': product_data.get('createdAt', datetime.now().isoformat()),
                        'originalUnfoundId': product_data.get('originalUnfoundId'),
                        'recentlyAddedId': product_data.get('recentlyAddedId'),
                        'scrapedAt': product_data.get('scrapedAt')
                    }
                    
                    # Add to barcode_cache collection
                    db.collection('barcode_cache').document(barcode).set(barcode_cache_data)
                    migrated_count += 1
                    print(f"Migrated product {barcode} to barcode_cache")
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully migrated {migrated_count} products to barcode_cache collection',
            'migratedCount': migrated_count
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/debug-verify', methods=['POST'])
def debug_verify():
    """Debug verification function"""
    try:
        data = request.get_json()
        product_ids = data.get('productIds', [])
        
        print(f"DEBUG: Received product IDs: {product_ids}", flush=True)
        
        if not db:
            print("DEBUG: Database not available", flush=True)
            return jsonify({'status': 'error', 'message': 'Database not available'}), 500
        
        print("DEBUG: Database is available", flush=True)
        
        # Test with first product ID
        if product_ids:
            product_id = product_ids[0]
            print(f"DEBUG: Testing with product ID: {product_id}", flush=True)
            
            product_doc = db.collection('recently_added_products').document(product_id).get()
            print(f"DEBUG: Product document exists: {product_doc.exists}", flush=True)
            
            if product_doc.exists:
                product_data = product_doc.to_dict()
                print(f"DEBUG: Product data keys: {list(product_data.keys())}", flush=True)
                print(f"DEBUG: Barcode: {product_data.get('barcode')}", flush=True)
                
                barcode = product_data.get('barcode')
                if barcode:
                    existing_product = db.collection('products').document(barcode).get()
                    print(f"DEBUG: Product exists in products collection: {existing_product.exists}", flush=True)
                    
                    # Try to manually add the product
                    test_product_data = {
                        'barcode': barcode,
                        'name': product_data.get('productName', 'Test Product'),
                        'price': 'N/A',
                        'image': None,
                        'source': 'debug_test',
                        'verified': True,
                        'verifiedAt': datetime.now().isoformat(),
                        'createdAt': datetime.now().isoformat()
                    }
                    
                    print(f"DEBUG: Attempting to add test product to barcode_cache: {test_product_data}", flush=True)
                    db.collection('barcode_cache').document(barcode).set(test_product_data)
                    print(f"DEBUG: Test product added to barcode_cache successfully", flush=True)
                    
                    # Verify it was added
                    added_product = db.collection('barcode_cache').document(barcode).get()
                    print(f"DEBUG: Verification - product exists in barcode_cache after adding: {added_product.exists}", flush=True)
        
        return jsonify({'status': 'success', 'message': 'Debug completed'})
    except Exception as e:
        print(f"DEBUG: Error in debug function: {e}", flush=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/recently-added-products/verify', methods=['POST'])
def verify_recently_added_products():
    """Mark recently added products as verified and move them to main products collection"""
    try:
        if not db:
            return jsonify({
                'status': 'error',
                'message': 'Database not available'
            }), 500
        
        data = request.get_json()
        product_ids = data.get('productIds', [])
        
        if not product_ids:
            return jsonify({
                'status': 'error',
                'message': 'No product IDs provided'
            }), 400
        
        verified_count = 0
        moved_to_products_count = 0
        
        for product_id in product_ids:
            try:
                print(f"DEBUG: Processing verification for product_id: {product_id}", flush=True)
                
                # Get the product data from recently_added_products
                product_doc = db.collection('recently_added_products').document(product_id).get()
                
                if not product_doc.exists:
                    print(f"DEBUG: Product {product_id} not found in recently_added_products")
                    continue
                
                product_data = product_doc.to_dict()
                barcode = product_data.get('barcode')
                
                print(f"DEBUG: Product data: {product_data}")
                print(f"DEBUG: Barcode: {barcode}")
                
                if not barcode:
                    print(f"DEBUG: No barcode found for product {product_id}")
                    continue
                
                # Check if product already exists in barcode_cache collection (where main app looks for products)
                existing_product = db.collection('barcode_cache').document(barcode).get()
                
                print(f"DEBUG: Checking if product {barcode} exists in barcode_cache collection: {existing_product.exists}")
                
                if existing_product.exists:
                    print(f"DEBUG: Product with barcode {barcode} already exists in barcode_cache collection")
                    # Remove from recently_added_products since it's already verified and in barcode_cache
                    db.collection('recently_added_products').document(product_id).delete()
                    verified_count += 1
                    print(f"✅ Removed verified product {barcode} from recently_added_products collection")
                    continue
                
                # Get scraped data if available
                scraped_data = product_data.get('scrapedData', {})
                
                # Create product data for barcode_cache collection (where main app looks for products)
                barcode_cache_data = {
                    'barcode': barcode,
                    'name': product_data.get('productName', scraped_data.get('name', 'Unknown')),
                    'price': product_data.get('price', scraped_data.get('price', 'N/A')),
                    'mrp': product_data.get('mrp', scraped_data.get('mrp', product_data.get('price', 'N/A'))),
                    'image': product_data.get('image', scraped_data.get('image')),
                    'brand': product_data.get('brand', scraped_data.get('brand', '')),
                    'category': product_data.get('category', scraped_data.get('category', '')),
                    'description': product_data.get('description', scraped_data.get('description', '')),
                    'source': 'background_processor_verified',
                    'verified': True,
                    'verifiedAt': datetime.now().isoformat(),
                    'createdAt': product_data.get('addedAt', datetime.now().isoformat()),
                    'originalUnfoundId': product_data.get('originalUnfoundId'),
                    'recentlyAddedId': product_id,
                    'scrapedAt': scraped_data.get('scrapedAt', product_data.get('addedAt'))
                }
                
                # Add to barcode_cache collection (where main app looks for products)
                print(f"DEBUG: Adding product to barcode_cache collection: {barcode_cache_data}")
                db.collection('barcode_cache').document(barcode).set(barcode_cache_data)
                moved_to_products_count += 1
                print(f"DEBUG: Successfully added product {barcode} to barcode_cache collection")
                
                # Remove from recently_added_products after successful verification and move
                db.collection('recently_added_products').document(product_id).delete()
                verified_count += 1
                
                print(f"✅ Verified, moved product {barcode} to barcode_cache collection, and removed from recently_added_products")
                
            except Exception as e:
                print(f"Error verifying product {product_id}: {e}")
                continue
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully verified {verified_count} products, moved {moved_to_products_count} to barcode_cache collection, and removed from recently added tab',
            'verifiedCount': verified_count,
            'movedToProductsCount': moved_to_products_count
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to verify products: {str(e)}'
        }), 500

@app.route('/api/products/unverified', methods=['GET'])
def get_unverified_products():
    """Get unverified products from main database for admin review"""
    try:
        if not db:
            return jsonify({
                'status': 'error',
                'message': 'Database not available'
            }), 500
        
        # Get unverified products from barcode_cache collection
        unverified_products = []
        barcode_cache_ref = db.collection('barcode_cache')
        
        # Filter for unverified products
        query = barcode_cache_ref.where('verified', '==', False)
        docs = query.stream()
        
        for doc in docs:
            product_data = doc.to_dict()
            product_data['id'] = doc.id
            unverified_products.append(product_data)
        
        # Sort by createdAt (newest first)
        unverified_products.sort(key=lambda x: x.get('createdAt', ''), reverse=True)
        
        return jsonify({
            'status': 'success',
            'data': unverified_products,
            'count': len(unverified_products)
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to get unverified products: {str(e)}'
        }), 500

@app.route('/api/products/bulk-verify', methods=['POST'])
def bulk_verify_products():
    """Bulk verify products by setting verified: true"""
    try:
        if not db:
            return jsonify({
                'status': 'error',
                'message': 'Database not available'
            }), 500
        
        data = request.get_json()
        product_ids = data.get('productIds', [])
        
        if not product_ids:
            return jsonify({
                'status': 'error',
                'message': 'Product IDs required'
            }), 400
        
        verified_count = 0
        current_time = datetime.now().isoformat()
        
        for product_id in product_ids:
            try:
                # Update product to verified: true
                db.collection('barcode_cache').document(product_id).update({
                    'verified': True,
                    'verifiedAt': current_time
                })
                verified_count += 1
                print(f"✅ Verified product {product_id}")
                
            except Exception as e:
                print(f"Error verifying product {product_id}: {e}")
                continue
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully verified {verified_count} out of {len(product_ids)} products',
            'verifiedCount': verified_count,
            'totalRequested': len(product_ids)
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to verify products: {str(e)}'
        }), 500

@app.route('/api/products/update-missing-images', methods=['POST'])
def update_missing_images():
    """Update products that have null/empty images with placeholder thumbnails"""
    try:
        if not db:
            return jsonify({'status': 'error', 'message': 'Database not available'}), 500
        
        updated_count = 0
        placeholder_image = "https://via.placeholder.com/300x300/cccccc/666666?text=Add+Image"
        
        # Get all products from barcode_cache collection
        barcode_cache_ref = db.collection('barcode_cache')
        docs = barcode_cache_ref.stream()
        
        for doc in docs:
            try:
                product_data = doc.to_dict()
                
                # Check if image is null, empty, or None
                if not product_data.get('image') or product_data.get('image') == 'null' or product_data.get('image') == '':
                    # Update with placeholder image
                    db.collection('barcode_cache').document(doc.id).update({
                        'image': placeholder_image,
                        'imageUpdatedAt': datetime.now().isoformat()
                    })
                    updated_count += 1
                    print(f"Updated product {doc.id} with placeholder image")
                    
            except Exception as e:
                print(f"Error updating product {doc.id}: {e}")
                continue
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully updated {updated_count} products with placeholder images',
            'updatedCount': updated_count
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/recently-added-products/update-fields', methods=['POST'])
def update_recently_added_products_fields():
    """Update existing recently added products with new fields (MRP, brand, category, description)"""
    try:
        if not db:
            return jsonify({'status': 'error', 'message': 'Database not available'}), 500
        
        updated_count = 0
        
        # Get all recently added products
        recently_added_ref = db.collection('recently_added_products')
        docs = recently_added_ref.stream()
        
        for doc in docs:
            try:
                product_data = doc.to_dict()
                
                # Check if MRP field is missing (indicating old format)
                if 'mrp' not in product_data:
                    # Update with new fields
                    update_data = {
                        'mrp': product_data.get('price', 'N/A'),  # Use price as MRP if not available
                        'brand': product_data.get('brand', ''),
                        'category': product_data.get('category', ''),
                        'description': product_data.get('description', ''),
                        'updatedAt': datetime.now().isoformat()
                    }
                    
                    # Update scrapedData as well
                    scraped_data = product_data.get('scrapedData', {})
                    if scraped_data:
                        scraped_data.update({
                            'mrp': product_data.get('price', 'N/A'),
                            'brand': product_data.get('brand', ''),
                            'category': product_data.get('category', ''),
                            'description': product_data.get('description', '')
                        })
                        update_data['scrapedData'] = scraped_data
                    
                    # Update the document
                    db.collection('recently_added_products').document(doc.id).update(update_data)
                    updated_count += 1
                    print(f"Updated recently added product {doc.id} with new fields")
                    
            except Exception as e:
                print(f"Error updating product {doc.id}: {e}")
                continue
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully updated {updated_count} recently added products with new fields',
            'updatedCount': updated_count
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/recently-added-products/clear', methods=['POST'])
def clear_recently_added_products():
    """Remove verified recently added products from the list"""
    try:
        if not db:
            return jsonify({
                'status': 'error',
                'message': 'Database not available'
            }), 500
        
        data = request.get_json()
        product_ids = data.get('productIds', [])
        
        if not product_ids:
            return jsonify({
                'status': 'error',
                'message': 'No product IDs provided'
            }), 400
        
        cleared_count = 0
        for product_id in product_ids:
            try:
                db.collection('recently_added_products').document(product_id).delete()
                cleared_count += 1
            except Exception as e:
                print(f"Error clearing product {product_id}: {e}")
                continue
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully cleared {cleared_count} products',
            'clearedCount': cleared_count
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to clear products: {str(e)}'
        }), 500

def process_unfound_barcodes_background():
    """Background job to process unfound barcodes"""
    global processing_status, processed_barcodes_history
    
    if not db:
        print("DEBUG: Background processor - Database not available")
        return
    
    try:
        processing_status['running'] = True
        processing_status['last_run'] = datetime.now().isoformat()
        processing_status['processed_count'] = 0
        processing_status['success_count'] = 0
        processing_status['error_count'] = 0
        
        print("DEBUG: Background processor started")
        
        # Get all unfound barcodes for processing
        unfound_barcodes_ref = db.collection('unfound_barcodes')
        unfound_barcodes = []
        
        for doc in unfound_barcodes_ref.stream():
            barcode_data = doc.to_dict()
            barcode_data['id'] = doc.id
            
            # Process all unfound barcodes - no retry logic needed since we delete failed ones
            unfound_barcodes.append(barcode_data)
            print(f"DEBUG: Adding barcode for processing: {barcode_data['barcode']}")
        
        print(f"DEBUG: Found {len(unfound_barcodes)} barcodes to process")
        
        for barcode_data in unfound_barcodes:
            try:
                processing_status['current_barcode'] = barcode_data['barcode']
                processing_status['processed_count'] += 1
                
                print(f"DEBUG: Processing barcode {barcode_data['barcode']} ({processing_status['processed_count']}/{len(unfound_barcodes)})")
                
                # Try to fetch product data
                url = f"https://smartconsumer-beta.org/01/{barcode_data['barcode']}"
                result = fetch_product_data_internal(barcode_data['barcode'], url)
                
                # Record processing result in history
                processed_at = datetime.now().isoformat()
                
                if result and result.get('success'):
                    # Product found! Add to products collection
                    product_data = result['product']
                    product_data['source'] = 'background_retry'
                    product_data['originalUnfoundId'] = barcode_data['id']
                    product_data['createdAt'] = processed_at
                    
                    # Add directly to barcode_cache collection (main database) with verified: false
                    barcode_cache_data = {
                        'barcode': barcode_data['barcode'],
                        'name': product_data.get('name', 'Unknown'),
                        'price': product_data.get('price', 'N/A'),
                        'mrp': product_data.get('mrp', product_data.get('price', 'N/A')),
                        'image': product_data.get('image'),
                        'brand': product_data.get('brand', ''),
                        'category': product_data.get('category', ''),
                        'description': product_data.get('description', ''),
                        'verified': False,  # Ready for admin verification
                        'source': 'background_processor',
                        'createdAt': processed_at,
                        'originalUnfoundId': barcode_data['id'],
                        'scrapedAt': processed_at
                    }
                    db.collection('barcode_cache').document(barcode_data['barcode']).set(barcode_cache_data)
                    
                    # Remove from unfound barcodes
                    db.collection('unfound_barcodes').document(barcode_data['id']).delete()
                    
                    processing_status['success_count'] += 1
                    print(f"DEBUG: ✅ Successfully found and added product: {product_data['name']}")
                    
                    # Add to processed history
                    processed_barcodes_history.append({
                        'barcode': barcode_data['barcode'],
                        'productName': product_data.get('name', 'Unknown'),
                        'success': True,
                        'processedAt': processed_at,
                        'result': f"Added: {product_data.get('name', 'Unknown')}"
                    })
                    
                else:
                    # Still not found, delete the barcode instead of retrying
                    db.collection('unfound_barcodes').document(barcode_data['id']).delete()
                    
                    processing_status['error_count'] += 1
                    print(f"DEBUG: ❌ Not found, deleting barcode: {barcode_data['barcode']}")
                    
                    # Add to processed history
                    processed_barcodes_history.append({
                        'barcode': barcode_data['barcode'],
                        'productName': None,
                        'success': False,
                        'processedAt': processed_at,
                        'result': 'Deleted - Not Found',
                        'error': 'Product not found on Smart Consumer - deleted from unfound list'
                    })
                
                # Keep only last 100 processed barcodes to prevent memory issues
                if len(processed_barcodes_history) > 100:
                    processed_barcodes_history = processed_barcodes_history[-100:]
                
                # Add human-like delay between requests to avoid being blocked
                import random
                delay = random.uniform(3, 6)  # Reduced delay between 3-6 seconds
                print(f"DEBUG: Waiting {delay:.1f} seconds before next barcode (human-like delay)...")
                time.sleep(delay)
                
            except Exception as e:
                processing_status['error_count'] += 1
                print(f"DEBUG: Error processing barcode {barcode_data['barcode']}: {e}")
                
                # Add error to processed history
                processed_barcodes_history.append({
                    'barcode': barcode_data['barcode'],
                    'productName': None,
                    'success': False,
                    'processedAt': datetime.now().isoformat(),
                    'result': 'Error',
                    'error': str(e)
                })
                continue
        
        print(f"DEBUG: Background processing completed. Processed: {processing_status['processed_count']}, Success: {processing_status['success_count']}, Errors: {processing_status['error_count']}")
        
    except Exception as e:
        print(f"DEBUG: Background processor error: {e}")
    finally:
        processing_status['running'] = False
        processing_status['current_barcode'] = None

def fetch_product_data_internal(barcode, url):
    """Internal function to fetch product data (used by background processor)"""
    driver = None
    try:
        # Setup Chrome options for headless browsing
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Initialize Chrome driver
        try:
            driver = webdriver.Chrome(options=chrome_options)
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        except Exception as e:
            print(f"DEBUG: Background processor - Chrome driver failed: {e}")
            return None
        
        # Navigate to URL
        driver.get(url)
        
        # Human-like delay after page load
        import random
        delay = random.uniform(1.5, 3)  # Random delay between 1.5-3 seconds
        print(f"DEBUG: Background processor - Waiting {delay:.1f} seconds (human-like delay)...")
        time.sleep(delay)
        
        # Simulate human-like behavior
        try:
            # Random scroll to simulate human reading
            scroll_amount = random.randint(50, 200)
            driver.execute_script(f"window.scrollBy(0, {scroll_amount});")
            time.sleep(random.uniform(0.3, 0.8))
        except Exception as e:
            print(f"DEBUG: Background processor - Error in human-like simulation: {e}")
        
        # Wait for page to load
        wait = WebDriverWait(driver, 10)
        try:
            wait.until(EC.any_of(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1")),
                EC.presence_of_element_located((By.CSS_SELECTOR, ".error"))
            ))
        except TimeoutException:
            pass
        
        # Extract product data
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        product_data = extract_product_data_selenium(driver, soup, barcode)
        
        if product_data and (product_data.get('name') != 'N/A' or product_data.get('price') != 'N/A'):
            return {'success': True, 'product': product_data}
        else:
            return {'success': False, 'product': create_empty_product_data(barcode)}
            
    except Exception as e:
        print(f"DEBUG: Background processor - Error fetching {barcode}: {e}")
        return None
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass

def start_background_processor():
    """Start the background processor in continuous real-time mode"""
    global background_processor
    
    if background_processor and background_processor.is_alive():
        print("DEBUG: Background processor already running")
        return False
    
    def continuous_processor():
        """Continuously process unfound barcodes in real-time"""
        processing_status['running'] = True
        print("DEBUG: 🚀 Background processor started in CONTINUOUS REAL-TIME mode")
        
        while processing_status['running']:
            try:
                # Check for unfound barcodes to process
                unfound_barcodes = get_unfound_barcodes_for_processing()
                
                if unfound_barcodes:
                    print(f"DEBUG: 🔄 Found {len(unfound_barcodes)} barcodes to process")
                    
                    # Process each barcode
                    for i, barcode_data in enumerate(unfound_barcodes):
                        if not processing_status['running']:
                            print("DEBUG: Background processor stopped by user")
                            break
                            
                        print(f"DEBUG: Processing barcode {barcode_data['barcode']} ({i+1}/{len(unfound_barcodes)})")
                        processing_status['current_barcode'] = barcode_data['barcode']
                        
                        # Process the barcode
                        result = process_single_barcode(barcode_data)
                        
                        if result:
                            print(f"DEBUG: ✅ Successfully processed {barcode_data['barcode']}")
                        else:
                            print(f"DEBUG: ❌ Failed to process {barcode_data['barcode']}")
                        
                        # Human-like delay between barcodes (2-5 seconds)
                        import random
                        delay = random.uniform(2, 5)
                        print(f"DEBUG: Waiting {delay:.1f} seconds before next barcode...")
                        time.sleep(delay)
                    
                    processing_status['current_barcode'] = None
                    print(f"DEBUG: ✅ Completed processing batch of {len(unfound_barcodes)} barcodes")
                else:
                    print("DEBUG: No unfound barcodes to process, waiting 30 seconds...")
                    time.sleep(30)  # Wait 30 seconds before checking again
                    
            except Exception as e:
                print(f"DEBUG: Error in continuous processor: {e}")
                time.sleep(60)  # Wait 1 minute on error before retrying
        
        print("DEBUG: Background processor stopped")
    
    background_processor = threading.Thread(target=continuous_processor, daemon=True)
    background_processor.start()
    
    print("DEBUG: 🚀 Background processor started in CONTINUOUS REAL-TIME mode")
    return True

def get_unfound_barcodes_for_processing():
    """Get unfound barcodes that need processing"""
    try:
        if not db:
            print("DEBUG: Database not available")
            return []
        
        unfound_barcodes = []
        unfound_ref = db.collection('unfound_barcodes')
        docs = unfound_ref.stream()
        
        for doc in docs:
            barcode_data = doc.to_dict()
            barcode_data['id'] = doc.id
            
            # Process all unfound barcodes - no retry logic needed since we delete failed ones
            unfound_barcodes.append(barcode_data)
            print(f"DEBUG: Adding barcode for processing: {barcode_data['barcode']}")
        
        print(f"DEBUG: Found {len(unfound_barcodes)} barcodes to process")
        return unfound_barcodes
        
    except Exception as e:
        print(f"DEBUG: Error getting unfound barcodes: {e}")
        return []

def process_single_barcode(barcode_data):
    """Process a single barcode and return success status"""
    try:
        barcode = barcode_data['barcode']
        print(f"DEBUG: Processing barcode: {barcode}")
        
        # Add human-like delay
        import random
        delay = random.uniform(2, 4)
        print(f"DEBUG: Background processor - Waiting {delay:.1f} seconds (human-like delay)...")
        time.sleep(delay)
        
        # Try to fetch product data
        url = f"https://smartconsumer-beta.org/{barcode}"
        result = fetch_product_data_internal(barcode, url)
        
        if result and result.get('status') == 'success':
            product_data = result['product']
            product_data['source'] = 'background_retry'
            product_data['originalUnfoundId'] = barcode_data['id']
            product_data['createdAt'] = datetime.now().isoformat()
            
            # Add directly to barcode_cache collection (main database) with verified: false
            barcode_cache_data = {
                'barcode': barcode,
                'name': product_data.get('name', 'Unknown'),
                'price': product_data.get('price', 'N/A'),
                'mrp': product_data.get('mrp', product_data.get('price', 'N/A')),
                'image': product_data.get('image'),
                'brand': product_data.get('brand', ''),
                'category': product_data.get('category', ''),
                'description': product_data.get('description', ''),
                'verified': False,  # Ready for admin verification
                'source': 'background_processor',
                'createdAt': datetime.now().isoformat(),
                'originalUnfoundId': barcode_data['id'],
                'scrapedAt': datetime.now().isoformat()
            }
            print(f"DEBUG: Adding to barcode_cache: {barcode_cache_data}")
            db.collection('barcode_cache').document(barcode).set(barcode_cache_data)
            print(f"DEBUG: ✅ Successfully added {barcode} to barcode_cache")
            
            # Remove from unfound barcodes
            print(f"DEBUG: Removing {barcode} from unfound_barcodes")
            db.collection('unfound_barcodes').document(barcode_data['id']).delete()
            print(f"DEBUG: ✅ Successfully removed {barcode} from unfound_barcodes")
            
            print(f"DEBUG: ✅ Successfully processed and moved {barcode} to main database")
            return True
        else:
            # Update retry count and timestamp
            retry_count = barcode_data.get('retryCount', 0) + 1
            db.collection('unfound_barcodes').document(barcode_data['id']).update({
                'retryCount': retry_count,
                'lastRetry': datetime.now().isoformat()
            })
            
            print(f"DEBUG: ❌ Still not found: {barcode} (retry #{retry_count})")
            
            # Human-like delay before next barcode
            delay = random.uniform(8, 15)
            print(f"DEBUG: Waiting {delay:.1f} seconds before next barcode (human-like delay)...")
            time.sleep(delay)
            
            return False
            
    except Exception as e:
        print(f"DEBUG: Error processing barcode {barcode_data['barcode']}: {e}")
        return False

def stop_background_processor():
    """Stop the background processor"""
    global background_processor
    processing_status['running'] = False
    processing_status['current_barcode'] = None
    print("DEBUG: Background processor stop requested")

def fallback_to_requests_scraping(url, barcode):
    """Fallback to requests-based scraping when Selenium fails"""
    try:
        print("DEBUG: Using requests-based scraping fallback...")
        
        # Enhanced headers to better mimic a real browser
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Cache-Control': 'max-age=0',
            'DNT': '1'
        }
        
        # Add a random delay to avoid being blocked
        import random
        time.sleep(random.uniform(1, 3))
        
        # Create a session for better connection handling
        session = requests.Session()
        session.headers.update(headers)
        
        # Make the request with retries
        max_retries = 3
        for attempt in range(max_retries):
            try:
                response = session.get(url, timeout=15)
                response.raise_for_status()
                break
            except requests.exceptions.RequestException as e:
                if attempt == max_retries - 1:
                    raise e
                time.sleep(2 ** attempt)  # Exponential backoff
        
        # Parse the HTML
        soup = BeautifulSoup(response.content, 'html.parser')
        
        print(f"DEBUG: Fallback - Page title: {soup.find('title').get_text() if soup.find('title') else 'No title found'}")
        print(f"DEBUG: Fallback - Response status: {response.status_code}")
        print(f"DEBUG: Fallback - Content length: {len(response.content)}")
        
        # Extract product information from the page
        product_data = extract_product_data(soup, barcode)
        
        if product_data and (product_data.get('name') != 'N/A' or product_data.get('price') != 'N/A'):
            return jsonify({
                'success': True,
                'status': 'found_fallback',
                'message': 'Product data found using fallback method (basic web scraping)',
                'product': product_data
            }), 200
        else:
            return jsonify({
                'success': False,
                'status': 'not_found',
                'message': 'Product information not found. This website may require JavaScript to load content.',
                'product': product_data
            }), 200
            
    except requests.exceptions.RequestException as e:
        return jsonify({
            'success': False,
            'status': 'error',
            'message': f'Network error in fallback: {str(e)}',
            'product': create_empty_product_data(barcode)
        }), 500
        
    except Exception as e:
        return jsonify({
            'success': False,
            'status': 'error',
            'message': f'Unexpected error in fallback: {str(e)}',
            'product': create_empty_product_data(barcode)
        }), 500

# Data Getter API Endpoint
@app.route('/api/fetch-product-data', methods=['POST'])
def fetch_product_data():
    driver = None
    try:
        data = request.get_json()
        if not data or 'barcode' not in data:
            return jsonify({'error': 'Barcode is required'}), 400

        barcode = data['barcode']
        
        # Construct the Smart Consumer URL
        url = f"https://smartconsumer-beta.org/01/{barcode}"
        
        print(f"DEBUG: Fetching product data for barcode: {barcode}")
        print(f"DEBUG: URL: {url}")
        
        # Setup Chrome options for headless browsing
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # Run in headless mode
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Initialize Chrome driver with better Windows compatibility
        try:
            # Try multiple approaches for Chrome driver initialization
            driver = None
            
            # Approach 1: Use ChromeDriverManager with specific version
            try:
                print("DEBUG: Attempting Chrome driver initialization with ChromeDriverManager...")
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)
                print("DEBUG: Chrome driver initialized successfully with ChromeDriverManager")
            except Exception as e1:
                print(f"DEBUG: ChromeDriverManager failed: {e1}")
                
                # Approach 2: Try without service (let Selenium find driver)
                try:
                    print("DEBUG: Attempting Chrome driver initialization without service...")
                    driver = webdriver.Chrome(options=chrome_options)
                    print("DEBUG: Chrome driver initialized successfully without service")
                except Exception as e2:
                    print(f"DEBUG: Direct Chrome initialization failed: {e2}")
                    
                    # Approach 3: Try Edge WebDriver as fallback
                    try:
                        print("DEBUG: Attempting Edge WebDriver as fallback...")
                        from selenium.webdriver.edge.service import Service as EdgeService
                        from selenium.webdriver.edge.options import Options as EdgeOptions
                        from webdriver_manager.microsoft import EdgeChromiumDriverManager
                        
                        edge_options = EdgeOptions()
                        edge_options.add_argument('--headless')
                        edge_options.add_argument('--no-sandbox')
                        edge_options.add_argument('--disable-dev-shm-usage')
                        edge_options.add_argument('--disable-gpu')
                        edge_options.add_argument('--window-size=1920,1080')
                        edge_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
                        
                        edge_service = EdgeService(EdgeChromiumDriverManager().install())
                        driver = webdriver.Edge(service=edge_service, options=edge_options)
                        print("DEBUG: Edge WebDriver initialized successfully")
                    except Exception as e3:
                        print(f"DEBUG: Edge WebDriver failed: {e3}")
                        
                        # Approach 4: Fallback to requests-based scraping
                        print("DEBUG: Falling back to requests-based scraping...")
                        return fallback_to_requests_scraping(url, barcode)
            
            if driver:
                # Execute script to remove webdriver property
                driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
                print("DEBUG: Browser driver initialized and configured successfully")
            else:
                raise Exception("No driver could be initialized")
                
        except Exception as e:
            print(f"DEBUG: Error initializing browser driver: {e}")
            return jsonify({
                'success': False,
                'status': 'error',
                'message': f'Failed to initialize browser: {str(e)}. Please ensure Chrome or Edge browser is installed.',
                'product': create_empty_product_data(barcode)
            }), 500
        
        # Navigate to the URL
        try:
            print("DEBUG: Navigating to URL...")
            driver.get(url)
            
            # Human-like delay after page load
            import random
            delay = random.uniform(2, 4)  # Random delay between 2-4 seconds
            print(f"DEBUG: Waiting {delay:.1f} seconds (human-like delay)...")
            time.sleep(delay)
            
            # Additional wait for images to load
            print(f"DEBUG: Waiting additional 3 seconds for images to load...")
            time.sleep(3)
            
            # Simulate human-like behavior
            try:
                # Random scroll to simulate human reading
                scroll_amount = random.randint(100, 500)
                driver.execute_script(f"window.scrollBy(0, {scroll_amount});")
                time.sleep(random.uniform(0.5, 1.5))
                
                # Scroll back up
                driver.execute_script(f"window.scrollBy(0, -{scroll_amount//2});")
                time.sleep(random.uniform(0.3, 0.8))
            except Exception as e:
                print(f"DEBUG: Error in human-like simulation: {e}")
            
            # Wait for the page to load and look for product data
            wait = WebDriverWait(driver, 15)
            
            # Wait for either product data to load or error message
            try:
                # Look for product information elements
                product_elements = wait.until(
                    EC.any_of(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "[data-testid*='product']")),
                        EC.presence_of_element_located((By.CSS_SELECTOR, ".product-info")),
                        EC.presence_of_element_located((By.CSS_SELECTOR, ".product-details")),
                        EC.presence_of_element_located((By.CSS_SELECTOR, "h1")),
                        EC.presence_of_element_located((By.CSS_SELECTOR, ".error")),
                        EC.presence_of_element_located((By.CSS_SELECTOR, ".not-found"))
                    )
                )
                print("DEBUG: Page elements loaded")
            except TimeoutException:
                print("DEBUG: Timeout waiting for page elements")
                # Try to get page source anyway
                pass
            
            # Get page source and parse with BeautifulSoup
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            print(f"DEBUG: Page title: {soup.find('title').get_text() if soup.find('title') else 'No title found'}")
            print(f"DEBUG: Page source length: {len(page_source)}")
            
            # Extract product data using Selenium and BeautifulSoup
            product_data = extract_product_data_selenium(driver, soup, barcode)
            
            if product_data and (product_data.get('name') != 'N/A' or product_data.get('price') != 'N/A'):
                print("DEBUG: ✅ Required fields extracted successfully!")
                print(f"DEBUG: 📦 Barcode: {product_data.get('barcode')}")
                print(f"DEBUG: 📝 Name: {product_data.get('name')}")
                print(f"DEBUG: 💰 Price: {product_data.get('price')}")
                print(f"DEBUG: 🖼️ Image: {product_data.get('image')}")
                return jsonify({
                    'success': True,
                    'status': 'found',
                    'product': product_data
                }), 200
            else:
                print("DEBUG: No product data found or error page detected")
                return jsonify({
                    'success': False,
                    'status': 'not_found',
                    'message': 'Product information not found on Smart Consumer website',
                    'product': create_empty_product_data(barcode)
                }), 200
                
        except WebDriverException as e:
            print(f"DEBUG: WebDriver error: {e}")
            return jsonify({
                'success': False,
                'status': 'error',
                'message': f'Browser error: {str(e)}',
                'product': create_empty_product_data(barcode)
            }), 500
            
    except Exception as e:
        print(f"DEBUG: Unexpected error: {e}")
        return jsonify({
            'success': False,
            'status': 'error',
            'message': f'Unexpected error: {str(e)}',
            'product': create_empty_product_data(data.get('barcode', 'Unknown') if 'data' in locals() else 'Unknown')
        }), 500
        
    finally:
        # Always close the driver
        if driver:
            try:
                driver.quit()
                print("DEBUG: Chrome driver closed")
            except Exception as e:
                print(f"DEBUG: Error closing driver: {e}")

def create_empty_product_data(barcode):
    """Create empty product data structure with only required fields"""
    return {
        'barcode': barcode,
        'name': 'N/A',
        'price': 'N/A',
        'image': "https://via.placeholder.com/300x300/cccccc/666666?text=Add+Image"  # Default placeholder
    }

def extract_product_data_selenium(driver, soup, barcode):
    """Extract product data using Selenium and BeautifulSoup"""
    try:
        product_data = create_empty_product_data(barcode)
        
        # Check if we're on an error page first
        current_url = driver.current_url
        page_title = driver.title
        
        # Check for 404 or error pages
        if ('404' in current_url or 
            'error' in current_url.lower() or 
            'not-found' in current_url.lower() or
            'error' in page_title.lower()):
            print(f"DEBUG: Detected error page - URL: {current_url}, Title: {page_title}")
            return None
        
        # Check for specific error messages in page content (very precise)
        # Only reject if we find clear error indicators, not just any mention of "not found"
        error_messages = [
            "string indices must be integers, not 'str'",
            "404 error",
            "page not found",
            "invalid barcode",
            "barcode not found",
            "no product data available",
            "error: product not found"  # Only reject if it's clearly an error message
        ]
        
        page_text = driver.page_source.lower()
        for error_msg in error_messages:
            if error_msg.lower() in page_text:
                print(f"DEBUG: Detected error message in page: '{error_msg}'")
                return None
        
        # Additional check: Only reject if we find "product not found" in a clear error context
        if "product not found" in page_text:
            # Check if it's in a clear error context (like error messages, alerts, etc.)
            error_contexts = [
                "error: product not found",
                "alert: product not found", 
                "message: product not found",
                "status: product not found"
            ]
            
            is_clear_error = any(context in page_text for context in error_contexts)
            if not is_clear_error:
                print(f"DEBUG: Found 'product not found' but not in clear error context - proceeding with extraction")
            else:
                print(f"DEBUG: Detected clear error context with 'product not found'")
                return None
        
        # Try to extract product name using multiple strategies
        try:
            # Strategy 1: Look for product name in various selectors
            name_selectors = [
                "h1",
                "[data-testid*='product-name']",
                ".product-name",
                ".product-title",
                ".product-info h1",
                ".product-info h2",
                ".product-details h1",
                ".product-details h2"
            ]
            
            for selector in name_selectors:
                try:
                    element = driver.find_element(By.CSS_SELECTOR, selector)
                    if element and element.text.strip():
                        name_text = element.text.strip()
                        # Check if the extracted text is an error message
                        if any(error_msg.lower() in name_text.lower() for error_msg in error_messages):
                            print(f"DEBUG: Skipping error message as product name: '{name_text}'")
                            continue
                        product_data['name'] = name_text
                        print(f"DEBUG: Found product name using selector '{selector}': {product_data['name']}")
                        break
                except NoSuchElementException:
                    continue
            
            # Strategy 2: Look in page title if no name found
            if product_data['name'] == 'N/A':
                title_element = soup.find('title')
                if title_element and title_element.get_text().strip():
                    title_text = title_element.get_text().strip()
                    # Clean up title to extract product name
                    if 'Smart Consumer' not in title_text and len(title_text) > 5:
                        product_data['name'] = title_text
                        print(f"DEBUG: Found product name from title: {product_data['name']}")
            
        except Exception as e:
            print(f"DEBUG: Error extracting product name: {e}")
        
        # Try to extract MRP/Price information
        try:
            # Look for MRP specifically first
            mrp_selectors = [
                "[data-testid*='mrp']",
                ".mrp",
                ".product-mrp",
                ".max-retail-price",
                ".retail-price",
                "[class*='mrp']",
                "[class*='retail']"
            ]
            
            for selector in mrp_selectors:
                try:
                    element = driver.find_element(By.CSS_SELECTOR, selector)
                    if element and element.text.strip():
                        price_text = element.text.strip()
                        if any(currency in price_text for currency in ['₹', 'Rs', '$', '€', '£']):
                            product_data['price'] = price_text
                            print(f"DEBUG: Found MRP using selector '{selector}': {product_data['price']}")
                            break
                except NoSuchElementException:
                    continue
            
            # If MRP not found, look for general price
            if product_data['price'] == 'N/A':
                price_selectors = [
                    "[data-testid*='price']",
                    ".price",
                    ".product-price",
                    ".cost",
                    ".amount",
                    ".selling-price"
                ]
                
                for selector in price_selectors:
                    try:
                        element = driver.find_element(By.CSS_SELECTOR, selector)
                        if element and element.text.strip():
                            price_text = element.text.strip()
                            if any(currency in price_text for currency in ['₹', 'Rs', '$', '€', '£']):
                                product_data['price'] = price_text
                                print(f"DEBUG: Found price using selector '{selector}': {product_data['price']}")
                                break
                    except NoSuchElementException:
                        continue
            
            # Fallback: Look for price patterns in text
            if product_data['price'] == 'N/A':
                try:
                    # Look for price patterns in the page source
                    import re
                    price_patterns = [
                        r'₹\s*[\d,]+\.?\d*',
                        r'Rs\s*[\d,]+\.?\d*',
                        r'\$\s*[\d,]+\.?\d*',
                        r'€\s*[\d,]+\.?\d*',
                        r'£\s*[\d,]+\.?\d*'
                    ]
                    
                    page_text = driver.page_source
                    for pattern in price_patterns:
                        matches = re.findall(pattern, page_text)
                        if matches:
                            product_data['price'] = matches[0]
                            print(f"DEBUG: Found price using regex pattern: {product_data['price']}")
                            break
                except Exception as e:
                    print(f"DEBUG: Error in price regex search: {e}")
            
            # Additional fallback: Look for any element containing currency symbols
            if product_data['price'] == 'N/A':
                try:
                    # Look for elements with currency symbols using XPath
                    currency_elements = driver.find_elements(By.XPATH, "//*[contains(text(), '₹') or contains(text(), 'Rs') or contains(text(), '$')]")
                    for elem in currency_elements:
                        text = elem.text.strip()
                        if any(currency in text for currency in ['₹', 'Rs', '$', '€', '£']) and len(text) < 20:  # Reasonable price length
                            product_data['price'] = text
                            print(f"DEBUG: Found price using XPath fallback: {product_data['price']}")
                            break
                except Exception as e:
                    print(f"DEBUG: Error in XPath price search: {e}")
                    
        except Exception as e:
            print(f"DEBUG: Error extracting price: {e}")
        
        # Brand and category extraction removed - not required
        
        # Try to extract product image URL with robust fallback
        try:
            print(f"DEBUG: Starting robust image extraction...")
            
            # Scroll to trigger lazy loading
            print(f"DEBUG: Scrolling to trigger lazy loading...")
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight/2);")
            time.sleep(2)
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)
            
            # First, let's see what images are actually on the page
            all_img_elements = driver.find_elements(By.TAG_NAME, "img")
            print(f"DEBUG: Found {len(all_img_elements)} total img elements on the page")
            
            for idx, img in enumerate(all_img_elements[:10]):  # Show first 10 images
                try:
                    src = img.get_attribute('src')
                    alt = img.get_attribute('alt')
                    print(f"DEBUG: Image {idx+1}: src='{src}', alt='{alt}'")
                except:
                    pass
            
            # Enhanced image selectors for better coverage
            img_selectors = [
                "img",  # Most basic selector - try first
                "[data-testid*='product-image']",
                "[data-testid*='image']",
                ".product-image img",
                ".product-photo img",
                ".product-img img",
                ".product-picture img",
                ".main-image img",
                ".hero-image img",
                ".featured-image img",
                "img[alt*='product']",
                "img[alt*='Product']",
                "img[src*='product']",
                "img[src*='Product']",
                "img[src*='gs1datakart']",  # GS1 DataKart API images
                "img[src*='api.gs1datakart.org']",  # GS1 DataKart API images
                "img[class*='product']",
                "img[class*='main']",
                "img[class*='hero']",
                "img[class*='featured']",
                ".image-container img",
                ".photo-container img",
                ".img-container img",
                "picture img",
                "figure img"
            ]
            
            print(f"DEBUG: Trying {len(img_selectors)} image selectors...")
            
            for i, selector in enumerate(img_selectors):
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    print(f"DEBUG: Selector {i+1}/{len(img_selectors)} '{selector}': found {len(elements)} elements")
                    
                    for j, element in enumerate(elements):
                        try:
                            img_src = element.get_attribute('src')
                            img_alt = element.get_attribute('alt')
                            img_class = element.get_attribute('class')
                            img_width = element.get_attribute('width')
                            img_height = element.get_attribute('height')
                            
                            print(f"DEBUG: Element {j+1}: src='{img_src}', alt='{img_alt}', class='{img_class}', size={img_width}x{img_height}")
                            
                            if img_src and img_src.strip():
                                # Skip data URLs, placeholder images, and logos
                                if (img_src.startswith('data:') or 
                                    'placeholder' in img_src.lower() or
                                    'logo' in img_src.lower() or
                                    'icon' in img_src.lower()):
                                    print(f"DEBUG: Skipping image (placeholder/logo): {img_src}")
                                    continue
                                
                                # Skip very small images (but allow 0 dimensions as they might be CSS-sized)
                                if (img_width and img_width != '0' and int(img_width) < 50) or \
                                   (img_height and img_height != '0' and int(img_height) < 50):
                                    print(f"DEBUG: Skipping image (too small): {img_src}")
                                    continue
                                
                                # Handle different URL formats
                                if img_src.startswith('http'):
                                    product_data['image'] = img_src
                                    print(f"DEBUG: ✅ Found image using selector '{selector}': {product_data['image']}")
                                    break
                                elif img_src.startswith('/'):
                                    product_data['image'] = f"https://smartconsumer-beta.org{img_src}"
                                    print(f"DEBUG: ✅ Found image using selector '{selector}': {product_data['image']}")
                                    break
                                elif img_src.startswith('//'):
                                    product_data['image'] = f"https:{img_src}"
                                    print(f"DEBUG: ✅ Found image using selector '{selector}': {product_data['image']}")
                                    break
                        except Exception as e:
                            print(f"DEBUG: Error processing element {j+1}: {e}")
                            continue
                    
                    if product_data['image']:
                        break
                        
                except Exception as e:
                    print(f"DEBUG: Error with selector '{selector}': {e}")
                    continue
            
            # Fallback: Look for any img tag with reasonable dimensions
            if product_data['image'] is None:
                try:
                    all_images = driver.find_elements(By.TAG_NAME, "img")
                    for img in all_images:
                        try:
                            img_src = img.get_attribute('src')
                            img_width = img.get_attribute('width')
                            img_height = img.get_attribute('height')
                            
                            # Look for images that are likely product images (not icons/logos)
                            if img_src and (
                                (img_width and int(img_width) > 100) or 
                                (img_height and int(img_height) > 100) or
                                'product' in img_src.lower() or
                                'item' in img_src.lower()
                            ):
                                if img_src.startswith('http'):
                                    product_data['image'] = img_src
                                    print(f"DEBUG: Found image using fallback method: {product_data['image']}")
                                    break
                                elif img_src.startswith('/'):
                                    product_data['image'] = f"https://smartconsumer-beta.org{img_src}"
                                    print(f"DEBUG: Found image using fallback method: {product_data['image']}")
                                    break
                        except (ValueError, TypeError):
                            continue
                except Exception as e:
                    print(f"DEBUG: Error in fallback image search: {e}")
            
            # If still no image found, try to construct GS1 DataKart URL based on barcode pattern
            if product_data['image'] is None:
                try:
                    barcode = product_data['barcode']
                    if barcode and len(barcode) >= 13:
                        # Extract first 9 digits for GS1 DataKart pattern
                        prefix = barcode[:9]
                        # Try to construct GS1 DataKart image URL
                        gs1_url = f"https://api.gs1datakart.org/files/render?file_key=product_upload/{prefix}/{barcode}/{barcode}_f.png"
                        product_data['image'] = gs1_url
                        print(f"DEBUG: 🖼️ Constructed GS1 DataKart URL: {gs1_url}")
                    else:
                        product_data['image'] = "https://via.placeholder.com/300x300/cccccc/666666?text=Add+Image"
                        print(f"DEBUG: 🖼️ No image found, using placeholder thumbnail")
                except Exception as e:
                    product_data['image'] = "https://via.placeholder.com/300x300/cccccc/666666?text=Add+Image"
                    print(f"DEBUG: 🖼️ Error constructing GS1 URL, using placeholder: {e}")
            else:
                print(f"DEBUG: 🖼️ Final image URL: {product_data['image']}")
                    
        except Exception as e:
            print(f"DEBUG: Error extracting image: {e}")
            # Set placeholder thumbnail on any error
            product_data['image'] = "https://via.placeholder.com/300x300/cccccc/666666?text=Add+Image"
            print(f"DEBUG: 🖼️ Error occurred, using placeholder thumbnail")
        
        # Description extraction removed - not required
        
        # Check if we found the key required fields
        key_fields = ['name', 'price']  # Barcode is always present, image is optional
        found_key_data = any(product_data[field] != 'N/A' for field in key_fields)
        
        if found_key_data:
            print(f"DEBUG: Successfully extracted required fields:")
            print(f"DEBUG: - Barcode: {product_data['barcode']}")
            print(f"DEBUG: - Name: {product_data['name']}")
            print(f"DEBUG: - Price: {product_data['price']}")
            print(f"DEBUG: - Image: {product_data['image']}")
            return product_data
        else:
            print("DEBUG: No key product data found (name or price)")
            return None
            
    except Exception as e:
        print(f"DEBUG: Error in extract_product_data_selenium: {e}")
        return None

def extract_product_data(soup, barcode):
    """Extract product data from Smart Consumer HTML (fallback method)"""
    try:
        product_data = create_empty_product_data(barcode)
        
        # Try to extract product name (this will need to be adjusted based on actual HTML structure)
        title_element = soup.find('title')
        if title_element:
            product_data['name'] = title_element.get_text().strip()
        
        # Look for product information in various common patterns
        # These selectors may need to be adjusted based on the actual Smart Consumer website structure
        
        # Try to find product name in h1 or h2 tags
        name_element = soup.find(['h1', 'h2'], class_=lambda x: x and ('product' in x.lower() or 'title' in x.lower()))
        if name_element:
            product_data['name'] = name_element.get_text().strip()
        
        # Try to find price information using string search
        price_texts = soup.find_all(string=lambda text: text and ('₹' in text or 'Rs' in text or '$' in text))
        if price_texts:
            product_data['price'] = price_texts[0].strip()
        
        # Try to find product image with robust fallback
        try:
            print(f"DEBUG: Fallback - Starting image extraction...")
            
            # Enhanced image selectors for fallback scraping
            img_selectors = [
                "img",  # Most basic selector
                "img[src*='product']",
                "img[src*='Product']",
                "img[src*='gs1datakart']",  # GS1 DataKart API images
                "img[src*='api.gs1datakart.org']",  # GS1 DataKart API images
                "img[alt*='product']",
                "img[alt*='Product']",
                "img[class*='product']",
                "img[class*='main']",
                "img[class*='hero']",
                "img[class*='featured']",
                ".product-image img",
                ".product-photo img",
                ".product-img img",
                ".main-image img",
                ".hero-image img",
                ".featured-image img"
            ]
            
            print(f"DEBUG: Fallback - Trying {len(img_selectors)} image selectors...")
            
            for i, selector in enumerate(img_selectors):
                try:
                    elements = soup.select(selector)
                    print(f"DEBUG: Fallback - Selector {i+1}/{len(img_selectors)} '{selector}': found {len(elements)} elements")
                    
                    for j, element in enumerate(elements):
                        try:
                            img_src = element.get('src')
                            img_alt = element.get('alt')
                            img_class = element.get('class')
                            img_width = element.get('width')
                            img_height = element.get('height')
                            
                            print(f"DEBUG: Fallback - Element {j+1}: src='{img_src}', alt='{img_alt}', class='{img_class}', size={img_width}x{img_height}")
                            
                            if img_src and img_src.strip():
                                # Skip data URLs, placeholder images, and logos
                                if (img_src.startswith('data:') or 
                                    'placeholder' in img_src.lower() or
                                    'logo' in img_src.lower() or
                                    'icon' in img_src.lower()):
                                    print(f"DEBUG: Fallback - Skipping image (placeholder/logo): {img_src}")
                                    continue
                                
                                # Skip very small images (but allow 0 dimensions as they might be CSS-sized)
                                if (img_width and img_width != '0' and int(img_width) < 50) or \
                                   (img_height and img_height != '0' and int(img_height) < 50):
                                    print(f"DEBUG: Fallback - Skipping image (too small): {img_src}")
                                    continue
                                
                                # Handle different URL formats
                                if img_src.startswith('http'):
                                    product_data['image'] = img_src
                                    print(f"DEBUG: Fallback - ✅ Found image using selector '{selector}': {product_data['image']}")
                                    break
                                elif img_src.startswith('/'):
                                    product_data['image'] = f"https://smartconsumer-beta.org{img_src}"
                                    print(f"DEBUG: Fallback - ✅ Found image using selector '{selector}': {product_data['image']}")
                                    break
                                elif img_src.startswith('//'):
                                    product_data['image'] = f"https:{img_src}"
                                    print(f"DEBUG: Fallback - ✅ Found image using selector '{selector}': {product_data['image']}")
                                    break
                        except Exception as e:
                            print(f"DEBUG: Fallback - Error processing element {j+1}: {e}")
                            continue
                    
                    if product_data['image']:
                        break
                        
                except Exception as e:
                    print(f"DEBUG: Fallback - Error with selector '{selector}': {e}")
                    continue
            
            # If still no image found, try to construct GS1 DataKart URL based on barcode pattern
            if product_data['image'] is None:
                try:
                    barcode = product_data['barcode']
                    if barcode and len(barcode) >= 13:
                        # Extract first 9 digits for GS1 DataKart pattern
                        prefix = barcode[:9]
                        # Try to construct GS1 DataKart image URL
                        gs1_url = f"https://api.gs1datakart.org/files/render?file_key=product_upload/{prefix}/{barcode}/{barcode}_f.png"
                        product_data['image'] = gs1_url
                        print(f"DEBUG: Fallback - 🖼️ Constructed GS1 DataKart URL: {gs1_url}")
                    else:
                        product_data['image'] = "https://via.placeholder.com/300x300/cccccc/666666?text=Add+Image"
                        print(f"DEBUG: Fallback - 🖼️ No image found, using placeholder thumbnail")
                except Exception as e:
                    product_data['image'] = "https://via.placeholder.com/300x300/cccccc/666666?text=Add+Image"
                    print(f"DEBUG: Fallback - 🖼️ Error constructing GS1 URL, using placeholder: {e}")
            else:
                print(f"DEBUG: Fallback - 🖼️ Final image URL: {product_data['image']}")
                    
        except Exception as e:
            print(f"DEBUG: Fallback - Error extracting image: {e}")
            # Set placeholder thumbnail on any error
            product_data['image'] = "https://via.placeholder.com/300x300/cccccc/666666?text=Add+Image"
            print(f"DEBUG: Fallback - 🖼️ Error occurred, using placeholder thumbnail")
        
        # If we found at least a name or price, return it
        if product_data['name'] != 'N/A' or product_data['price'] != 'N/A':
            return product_data
        else:
            return None
            
    except Exception as e:
        print(f"Error extracting product data: {e}")
        return None

@app.route('/api/import-barcodes', methods=['POST'])
@login_required
def import_barcodes_with_scraping():
    """Import barcodes from Excel and add them to unfound barcodes for background processing"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'File must be an Excel file (.xlsx or .xls)'}), 400
        
        # Read Excel file
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        
        # Get headers from first row
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
        
        # Check if barcode column exists (case insensitive)
        barcode_col_index = None
        for i, header in enumerate(headers):
            if 'barcode' in header.lower():
                barcode_col_index = i + 1  # Excel columns are 1-indexed
                break
        
        if barcode_col_index is None:
            return jsonify({'error': 'Excel file must contain a "barcode" column'}), 400
        
        print(f"Found barcode column at index {barcode_col_index}")
        
        # Process barcodes
        processed_count = 0
        added_to_unfound_count = 0
        skipped_count = 0
        skipped_already_scraped = 0
        skipped_already_unfound = 0
        errors = []
        
        print(f"Starting barcode import to unfound list...")
        print(f"Total rows to process: {ws.max_row - 1}")
        print(f"Headers found: {headers}")
        
        for row_num in range(2, ws.max_row + 1):
            try:
                # Get barcode value directly from the barcode column
                barcode_cell = ws.cell(row=row_num, column=barcode_col_index)
                barcode_value = barcode_cell.value
                
                # Handle different data types (string, int, float)
                if barcode_value is None:
                    print(f"Row {row_num}: Empty barcode cell, skipping")
                    continue
                
                # Convert to string and clean up
                barcode = str(barcode_value).strip()
                if not barcode or barcode.lower() in ['none', 'null', '']:
                    continue
                
                # Remove any non-numeric characters if needed (optional)
                # barcode = ''.join(filter(str.isdigit, barcode))
                
                print(f"Processing barcode: {barcode} (row {row_num})")
                
                # Check if barcode already exists in barcode_cache (successfully scraped)
                existing_doc = db.collection('barcode_cache').document(barcode).get()
                if existing_doc.exists:
                    print(f"Barcode {barcode} already exists in barcode_cache (successfully scraped), skipping")
                    skipped_count += 1
                    skipped_already_scraped += 1
                    continue
                
                # Check if barcode already exists in unfound_barcodes and update it instead of skipping
                unfound_query = db.collection('unfound_barcodes').where('barcode', '==', barcode).limit(1).get()
                if unfound_query:
                    # Update existing unfound barcode to reset retry count and timestamp
                    unfound_doc = unfound_query[0]
                    unfound_doc.reference.update({
                        'source': 'excel',
                        'createdAt': datetime.now().isoformat(),
                        'lastRetry': None,
                        'retryCount': 0,
                        'status': 'pending'
                    })
                    print(f"🔄 Updated existing unfound barcode {barcode} (reset retry count)")
                    added_to_unfound_count += 1
                    processed_count += 1
                    continue
                
                # Add to unfound_barcodes with source label
                unfound_data = {
                    'barcode': barcode,
                    'source': 'excel',
                    'createdAt': datetime.now().isoformat(),
                    'lastRetry': None,
                    'retryCount': 0,
                    'status': 'pending'
                }
                db.collection('unfound_barcodes').add(unfound_data)
                added_to_unfound_count += 1
                print(f"✅ Added barcode {barcode} to unfound list (source: excel)")
                
                processed_count += 1
                
            except Exception as e:
                error_msg = f"Row {row_num}: {str(e)}"
                errors.append(error_msg)
                print(f"Error processing row {row_num}: {e}")
        
        response = {
            'status': 'success',
            'message': f'Import completed successfully!',
            'processed_count': processed_count,
            'added_to_unfound_count': added_to_unfound_count,
            'skipped_count': skipped_count,
            'skipped_already_scraped': skipped_already_scraped,
            'skipped_already_unfound': skipped_already_unfound,
            'errors': errors[:10] if errors else []
        }
        
        print(f"Import completed: {processed_count} processed, {added_to_unfound_count} added to unfound, {skipped_count} skipped")
        return jsonify(response)
        
    except Exception as e:
        print(f"Import error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/clear-barcode-cache', methods=['POST'])
@login_required
def clear_barcode_cache():
    """Clear all barcodes from barcode_cache collection"""
    try:
        if not db:
            return jsonify({'error': 'Database not available'}), 500
        
        # Get all documents from barcode_cache collection
        barcode_cache_ref = db.collection('barcode_cache')
        docs = barcode_cache_ref.stream()
        
        cleared_count = 0
        for doc in docs:
            doc.reference.delete()
            cleared_count += 1
        
        print(f"Cleared {cleared_count} barcodes from cache")
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully cleared {cleared_count} barcodes from cache',
            'cleared_count': cleared_count
        })
        
    except Exception as e:
        print(f"Error clearing barcode cache: {e}")
        return jsonify({'error': str(e)}), 500

def scrape_product_data_for_import(barcode, url):
    """Scrape product data from Smart Consumer website for import"""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException
    from bs4 import BeautifulSoup
    
    driver = None
    try:
        # Setup Chrome options for headless browsing
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        # Initialize Chrome driver
        driver = webdriver.Chrome(options=chrome_options)
        driver.get(url)
        
        # Wait for page to load
        wait = WebDriverWait(driver, 10)
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "h1")))
        except TimeoutException:
            pass
        
        # Extract product data
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        
        # Extract product information
        product_data = {
            'name': 'N/A',
            'brand': '',
            'category': '',
            'mrp': 0.0,
            'salePrice': 0.0,
            'imageUrl': '',
            'description': '',
            'size': '',
            'unit': '',
            'isActive': True,
            'useInFirstStart': False
        }
        
        # Try to extract product name
        try:
            name_element = soup.find('h1')
            if name_element:
                product_data['name'] = name_element.get_text().strip()
        except:
            pass
        
        # Try to extract price
        try:
            price_elements = soup.find_all(['span', 'div'], class_=lambda x: x and 'price' in x.lower())
            for element in price_elements:
                text = element.get_text().strip()
                if '₹' in text or 'Rs' in text:
                    # Extract numeric value
                    import re
                    price_match = re.search(r'[\d,]+\.?\d*', text.replace(',', ''))
                    if price_match:
                        price = float(price_match.group())
                        product_data['mrp'] = price
                        product_data['salePrice'] = price
                        break
        except:
            pass
        
        # Try to extract image
        try:
            img_element = soup.find('img')
            if img_element and img_element.get('src'):
                product_data['imageUrl'] = img_element['src']
                product_data['photoPath'] = img_element['src']
        except:
            pass
        
        # Try to extract brand
        try:
            brand_elements = soup.find_all(['span', 'div'], class_=lambda x: x and 'brand' in x.lower())
            for element in brand_elements:
                text = element.get_text().strip()
                if text and len(text) < 50:  # Reasonable brand name length
                    product_data['brand'] = text
                    break
        except:
            pass
        
        return product_data
        
    except Exception as e:
        print(f"Error scraping {barcode}: {e}")
        return None
    finally:
        if driver:
            driver.quit()

if __name__ == '__main__':
    import os
    
    # Setup production logging
    def setup_production_logging():
        """Setup production logging configuration"""
        if not app.debug and not app.testing:
            import logging
            from logging.handlers import RotatingFileHandler
            
            # Create logs directory if it doesn't exist
            if not os.path.exists('logs'):
                os.makedirs('logs', exist_ok=True)
            
            # File handler for application logs
            file_handler = RotatingFileHandler(
                'logs/app.log', maxBytes=10240000, backupCount=10
            )
            file_handler.setFormatter(logging.Formatter(
                '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
            ))
            file_handler.setLevel(logging.INFO)
            app.logger.addHandler(file_handler)
            
            # Error handler for error logs
            error_handler = RotatingFileHandler(
                'logs/error.log', maxBytes=10240000, backupCount=10
            )
            error_handler.setFormatter(logging.Formatter(
                '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
            ))
            error_handler.setLevel(logging.ERROR)
            app.logger.addHandler(error_handler)
            
            app.logger.setLevel(logging.INFO)
            app.logger.info('EasyBill Admin Dashboard startup')
    
    setup_production_logging()
    
    print(f"Firebase Status: {firebase_status}")
    if db:
        print("Firebase connected successfully!")
    else:
        print("Firebase connection failed - dashboard will show connection status")
    
    # Get port from environment variable (for Render) or default to 5000
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    
    print(f"Dashboard available at: http://localhost:{port}")
    print(f"Environment: {os.environ.get('FLASK_ENV', 'development')}")
    print(f"Debug mode: {debug_mode}")
    
    app.run(debug=debug_mode, host='0.0.0.0', port=port)